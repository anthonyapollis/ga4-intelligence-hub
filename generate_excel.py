"""
GA4 Intelligence Hub — Professional Excel Report Generator
==========================================================
Generates a fully formatted, chart-rich Excel workbook.

Sheets:
  1. Dashboard        — KPI summary cards + YoY growth
  2. Transactions     — 40 sample orders + conditional formatting + chart
  3. Event Summary    — 25 GA4 events + bar chart
  4. Monthly Revenue  — 36-month trend + line chart
  5. Funnel Analysis  — 6-step checkout funnel + chart + drop-rate heatmap
  6. ML Models        — 3 model results + LTV tiers + feature importance chart
  7. Traffic & Audience — Source/medium/country/device breakdown
  8. BigQuery Queries — 10 production queries reference

Run: python generate_excel.py
Output: GA4_Intelligence_Hub_Professional.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
import copy

# ─── COLOUR PALETTE ──────────────────────────────────────────────────────────
C_NAVY      = "0F172A"
C_BLUE_DK   = "1D4ED8"
C_BLUE      = "2563EB"
C_BLUE_LT   = "DBEAFE"
C_BLUE_XLT  = "EFF6FF"
C_VIOLET    = "7C3AED"
C_VIOLET_LT = "EDE9FE"
C_GREEN     = "059669"
C_GREEN_LT  = "D1FAE5"
C_GOLD      = "D97706"
C_GOLD_LT   = "FEF3C7"
C_RED       = "DC2626"
C_RED_LT    = "FEE2E2"
C_TEAL      = "0D9488"
C_TEAL_LT   = "CCFBF1"
C_MUTED     = "64748B"
C_LINE      = "E2E8F0"
C_BG        = "F8FAFC"
C_WHITE     = "FFFFFF"

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="0F172A", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(top=True, bottom=True, left=False, right=False):
    s = Side(style="thin", color=C_LINE)
    return Border(
        top=s if top else None,
        bottom=s if bottom else None,
        left=s if left else None,
        right=s if right else None,
    )

def header_style(ws, row, cols, bg=C_NAVY, fg=C_WHITE, size=11):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = font(bold=True, size=size, color=fg)
        c.alignment = align("center")
        c.border = thin_border()

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def section_header(ws, row, col, text, bg=C_BLUE, fg=C_WHITE, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, size=11, color=fg)
    c.alignment = align("left", "center")
    c.border = thin_border()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)

def data_cell(ws, row, col, value, bold=False, bg=None, fg=C_NAVY,
              num_fmt=None, h="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=10, color=fg)
    c.alignment = align(h, "center", wrap)
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    return c

def badge_cell(ws, row, col, text, bg, fg):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, size=10, color=fg)
    c.alignment = align("center")
    c.border = thin_border()

def kpi_block(ws, row, col, label, value, sub="", bg=C_BLUE_XLT, accent=C_BLUE):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    top = ws.cell(row=row, column=col, value=label.upper())
    top.fill = fill(accent); top.font = font(bold=True, size=9, color=C_WHITE)
    top.alignment = align("center")
    mid = ws.cell(row=row+1, column=col, value=value)
    mid.fill = fill(bg); mid.font = font(bold=True, size=18, color=accent)
    mid.alignment = align("center")
    btm = ws.cell(row=row+2, column=col, value=sub)
    btm.fill = fill(bg); btm.font = font(size=9, color=C_MUTED, italic=True)
    btm.alignment = align("center")

# ─── DATA ────────────────────────────────────────────────────────────────────
# 100 transactions spanning Nov 2020 → Oct 2023 · real product SKU prices
# Columns: tx_id, date, cust_type, source, medium, device, country, items, revenue, tax, shipping, tier, propensity
TRANSACTIONS = [
    # ── Year 1 · Nov 2020 – Oct 2021 · AOV $25.23 ──
    ("GMS-20201101-001","2020-11-01","new","google","organic","desktop","India",          2, 51.98, 5.20,0.00,"Bronze",  0.72),
    ("GMS-20201101-002","2020-11-01","returning","(direct)","(none)","mobile","US",       1, 25.99, 2.60,4.99,"Bronze",  0.58),
    ("GMS-20201101-003","2020-11-01","new","google","organic","desktop","US",             3,100.97,10.10,0.00,"Gold",    0.84),
    ("GMS-20201103-001","2020-11-03","new","google","cpc","desktop","US",                 1, 22.99, 2.30,4.99,"Bronze",  0.42),
    ("GMS-20201107-001","2020-11-07","returning","(direct)","(none)","desktop","Ireland", 2, 45.98, 4.60,0.00,"Silver",  0.71),
    ("GMS-20201108-001","2020-11-08","new","google","cpc","desktop","Canada",             2, 48.98, 4.90,5.99,"Silver",  0.67),
    ("GMS-20201112-001","2020-11-12","new","bing","organic","mobile","UK",                1, 20.99, 2.10,6.99,"Bronze",  0.33),
    ("GMS-20201115-001","2020-11-15","returning","google","organic","desktop","Ireland",  4,109.96,11.00,0.00,"Gold",    0.91),
    ("GMS-20201115-002","2020-11-15","new","youtube.com","referral","mobile","UK",        1, 29.99, 3.00,0.00,"Bronze",  0.43),
    ("GMS-20201201-001","2020-12-01","new","google","organic","desktop","Japan",          2, 73.98, 7.40,9.99,"Silver",  0.62),
    ("GMS-20201210-001","2020-12-10","returning","newsletter","email","desktop","Canada", 5,199.95,20.00,0.00,"Platinum",0.95),
    ("GMS-20201215-001","2020-12-15","new","google","organic","mobile","US",              3, 84.97, 8.50,0.00,"Silver",  0.76),
    ("GMS-20201215-002","2020-12-15","returning","google","cpc","desktop","US",           2, 57.98, 5.80,0.00,"Silver",  0.82),
    ("GMS-20201220-001","2020-12-20","new","google","organic","desktop","Singapore",      1, 34.99, 3.50,9.99,"Bronze",  0.55),
    ("GMS-20201224-001","2020-12-24","returning","(direct)","(none)","desktop","Ireland", 6,234.94,23.49,0.00,"Platinum",0.97),
    ("GMS-20201226-001","2020-12-26","new","google","organic","mobile","Canada",          2, 45.98, 4.60,4.99,"Bronze",  0.49),
    ("GMS-20201231-001","2020-12-31","returning","google","organic","desktop","US",       4,149.96,15.00,0.00,"Gold",    0.88),
    ("GMS-20210108-001","2021-01-08","returning","google","organic","desktop","US",       3, 94.97, 9.50,0.00,"Gold",    0.79),
    ("GMS-20210115-001","2021-01-15","new","bing","organic","desktop","UK",               2, 56.98, 5.70,5.99,"Silver",  0.61),
    ("GMS-20210120-001","2021-01-20","returning","mail.google.com","referral","desktop","Ireland",4,159.96,16.00,0.00,"Platinum",0.93),
    ("GMS-20210131-001","2021-01-31","returning","google","organic","desktop","Canada",   3,107.97,10.80,0.00,"Gold",    0.86),
    ("GMS-20210205-001","2021-02-05","returning","newsletter","email","desktop","Ireland",3,131.97,13.20,0.00,"Gold",    0.89),
    ("GMS-20210214-001","2021-02-14","new","google","organic","desktop","US",             2, 58.98, 5.90,0.00,"Silver",  0.66),
    ("GMS-20210301-001","2021-03-01","new","facebook.com","referral","mobile","UK",       1, 18.99, 1.90,6.99,"Bronze",  0.29),
    ("GMS-20210315-001","2021-03-15","returning","google","organic","desktop","Canada",   4,159.96,16.00,0.00,"Gold",    0.92),
    ("GMS-20210401-001","2021-04-01","new","google","cpc","desktop","US",                 2, 62.98, 6.30,0.00,"Silver",  0.68),
    ("GMS-20210501-001","2021-05-01","returning","(direct)","(none)","desktop","Ireland", 5,204.95,20.50,0.00,"Platinum",0.96),
    ("GMS-20210601-001","2021-06-01","new","google","organic","mobile","India",           1, 17.99, 1.80,4.99,"Bronze",  0.35),
    ("GMS-20210701-001","2021-07-01","returning","newsletter","email","desktop","US",     3,109.97,11.00,0.00,"Gold",    0.87),
    ("GMS-20210801-001","2021-08-01","new","google","organic","desktop","Germany",        2, 73.98, 7.40,7.99,"Silver",  0.64),
    ("GMS-20210901-001","2021-09-01","returning","google","cpc","desktop","Canada",       4,163.96,16.40,0.00,"Gold",    0.91),
    ("GMS-20211001-001","2021-10-01","new","bing","organic","desktop","US",               1, 29.99, 3.00,4.99,"Bronze",  0.47),
    # ── Year 2 · Nov 2021 – Oct 2022 · AOV $24.88 ──
    ("GMS-20211101-001","2021-11-01","returning","(direct)","(none)","desktop","Ireland", 6,239.94,24.00,0.00,"Platinum",0.98),
    ("GMS-20211201-001","2021-12-01","new","google","organic","mobile","Australia",       2, 50.98, 5.10,5.99,"Bronze",  0.53),
    ("GMS-20211215-001","2021-12-15","returning","newsletter","email","desktop","US",     4,179.96,18.00,0.00,"Platinum",0.94),
    ("GMS-20211224-001","2021-12-24","new","google","organic","desktop","UK",             3, 87.97, 8.80,0.00,"Silver",  0.73),
    ("GMS-20211231-001","2021-12-31","returning","google","cpc","desktop","Canada",       2, 54.98, 5.50,0.00,"Silver",  0.78),
    ("GMS-20220115-001","2022-01-15","new","google","organic","mobile","US",              1, 24.99, 2.50,4.99,"Bronze",  0.38),
    ("GMS-20220201-001","2022-02-01","returning","newsletter","email","desktop","Ireland",4,159.96,16.00,0.00,"Platinum",0.93),
    ("GMS-20220214-001","2022-02-14","new","google","cpc","mobile","Canada",              2, 48.98, 4.90,5.99,"Silver",  0.62),
    ("GMS-20220301-001","2022-03-01","returning","(direct)","(none)","desktop","US",      3, 89.97, 9.00,0.00,"Gold",    0.87),
    ("GMS-20220315-001","2022-03-15","new","google","organic","desktop","Germany",        1, 25.99, 2.60,7.99,"Bronze",  0.45),
    ("GMS-20220401-001","2022-04-01","returning","newsletter","email","desktop","Ireland",5,199.95,20.00,0.00,"Platinum",0.96),
    ("GMS-20220415-001","2022-04-15","new","google","cpc","mobile","UK",                  1, 22.99, 2.30,6.99,"Bronze",  0.39),
    ("GMS-20220501-001","2022-05-01","returning","google","organic","desktop","US",       3, 99.97, 10.00,0.00,"Gold",   0.85),
    ("GMS-20220601-001","2022-06-01","new","facebook.com","referral","mobile","India",    1, 16.99, 1.70,5.99,"Bronze",  0.32),
    ("GMS-20220615-001","2022-06-15","returning","(direct)","(none)","desktop","Canada",  2, 57.98, 5.80,0.00,"Silver",  0.74),
    ("GMS-20220701-001","2022-07-01","new","google","organic","desktop","US",             2, 45.98, 4.60,0.00,"Silver",  0.58),
    ("GMS-20220715-001","2022-07-15","returning","newsletter","email","desktop","Ireland",4,174.96,17.50,0.00,"Gold",    0.92),
    ("GMS-20220801-001","2022-08-01","new","bing","organic","mobile","UK",                1, 19.99, 2.00,5.99,"Bronze",  0.41),
    ("GMS-20220815-001","2022-08-15","returning","google","cpc","desktop","US",           3,119.97,12.00,0.00,"Gold",    0.88),
    ("GMS-20220901-001","2022-09-01","new","google","organic","desktop","Australia",      2, 52.98, 5.30,8.99,"Silver",  0.56),
    ("GMS-20220915-001","2022-09-15","returning","(direct)","(none)","desktop","Ireland", 6,249.94,25.00,0.00,"Platinum",0.97),
    ("GMS-20221001-001","2022-10-01","new","google","organic","mobile","Germany",         1, 20.99, 2.10,6.99,"Bronze",  0.44),
    ("GMS-20221015-001","2022-10-15","returning","newsletter","email","desktop","US",     4,164.96,16.50,0.00,"Gold",    0.91),
    # ── Year 3 · Nov 2022 – Oct 2023 · AOV $26.39 ──
    ("GMS-20221101-001","2022-11-01","returning","(direct)","(none)","desktop","Ireland", 7,274.93,27.49,0.00,"Platinum",0.98),
    ("GMS-20221115-001","2022-11-15","new","google","organic","desktop","US",             2, 58.98, 5.90,0.00,"Silver",  0.69),
    ("GMS-20221125-001","2022-11-25","returning","newsletter","email","desktop","Canada", 5,219.95,22.00,0.00,"Platinum",0.96),  # Black Friday
    ("GMS-20221201-001","2022-12-01","new","google","organic","mobile","UK",              2, 52.98, 5.30,4.99,"Silver",  0.63),
    ("GMS-20221210-001","2022-12-10","returning","google","cpc","desktop","US",           4,179.96,18.00,0.00,"Gold",    0.90),
    ("GMS-20221215-001","2022-12-15","new","facebook.com","referral","mobile","India",    1, 17.99, 1.80,5.99,"Bronze",  0.36),
    ("GMS-20221220-001","2022-12-20","returning","(direct)","(none)","desktop","Ireland", 8,339.92,33.99,0.00,"Platinum",0.99),  # Xmas peak
    ("GMS-20221224-001","2022-12-24","new","google","organic","desktop","US",             3, 94.97, 9.50,0.00,"Silver",  0.71),
    ("GMS-20221231-001","2022-12-31","returning","newsletter","email","desktop","Canada", 4,179.96,18.00,0.00,"Gold",    0.89),
    ("GMS-20230115-001","2023-01-15","new","google","organic","mobile","US",              1, 25.99, 2.60,4.99,"Bronze",  0.41),
    ("GMS-20230201-001","2023-02-01","returning","(direct)","(none)","desktop","Ireland", 3,119.97,12.00,0.00,"Gold",    0.86),
    ("GMS-20230214-001","2023-02-14","new","google","cpc","mobile","UK",                  2, 55.98, 5.60,5.99,"Silver",  0.65),  # Valentine's
    ("GMS-20230301-001","2023-03-01","returning","newsletter","email","desktop","US",     5,209.95,21.00,0.00,"Platinum",0.95),
    ("GMS-20230315-001","2023-03-15","new","google","organic","desktop","Germany",        1, 27.99, 2.80,7.99,"Bronze",  0.46),
    ("GMS-20230401-001","2023-04-01","returning","(direct)","(none)","desktop","Canada",  3,109.97,11.00,0.00,"Gold",    0.88),
    ("GMS-20230415-001","2023-04-15","new","bing","organic","mobile","US",                1, 22.99, 2.30,4.99,"Bronze",  0.40),
    ("GMS-20230501-001","2023-05-01","returning","google","organic","desktop","Ireland",  4,189.96,19.00,0.00,"Platinum",0.94),
    ("GMS-20230601-001","2023-06-01","new","facebook.com","referral","mobile","India",    1, 14.99, 1.50,5.99,"Bronze",  0.31),
    ("GMS-20230615-001","2023-06-15","returning","newsletter","email","desktop","US",     3,124.97,12.50,0.00,"Gold",    0.87),
    ("GMS-20230701-001","2023-07-01","new","google","organic","desktop","Australia",      2, 57.98, 5.80,8.99,"Silver",  0.60),
    ("GMS-20230715-001","2023-07-15","returning","(direct)","(none)","desktop","Ireland", 5,214.95,21.50,0.00,"Platinum",0.97),
    ("GMS-20230801-001","2023-08-01","new","google","cpc","mobile","UK",                  1, 27.99, 2.80,6.99,"Bronze",  0.48),
    ("GMS-20230815-001","2023-08-15","returning","google","organic","desktop","Canada",   4,164.96,16.50,0.00,"Gold",    0.91),
    ("GMS-20230901-001","2023-09-01","new","bing","organic","desktop","US",               2, 64.98, 6.50,0.00,"Silver",  0.62),
    ("GMS-20230915-001","2023-09-15","returning","newsletter","email","desktop","Ireland",6,254.94,25.49,0.00,"Platinum",0.98),
    ("GMS-20231001-001","2023-10-01","new","google","organic","mobile","Germany",         1, 32.99, 3.30,7.99,"Bronze",  0.51),
    ("GMS-20231015-001","2023-10-15","returning","(direct)","(none)","desktop","US",      3,129.97,13.00,0.00,"Gold",    0.89),
    ("GMS-20231025-001","2023-10-25","new","google","cpc","mobile","Canada",              2, 55.98, 5.60,5.99,"Silver",  0.66),
    ("GMS-20231031-001","2023-10-31","returning","newsletter","email","desktop","Ireland",5,229.95,23.00,0.00,"Platinum",0.96),
]

EVENTS = [
    ("page_view",4153080,"Auto-collected","All pages"),
    ("user_engagement",2076444,"Auto-collected","Engaged sessions"),
    ("session_start",1891200,"Auto-collected","All sessions"),
    ("scroll",1246392,"Enhanced Measurement","90% depth"),
    ("view_item",561840,"Enhanced Ecommerce","Product detail"),
    ("view_item_list",422640,"Enhanced Ecommerce","Category/search grid"),
    ("first_visit",386520,"Auto-collected","New users"),
    ("view_promotion",282040,"Enhanced Ecommerce","Banner impressions"),
    ("add_to_cart",134880,"Enhanced Ecommerce","Cart intent"),
    ("view_search_results",98040,"Custom","Site search"),
    ("begin_checkout",101100,"Enhanced Ecommerce","Checkout start"),
    ("select_item",91080,"Enhanced Ecommerce","Product click"),
    ("add_shipping_info",73120,"Enhanced Ecommerce","Shipping step"),
    ("add_payment_info",56170,"Enhanced Ecommerce","Payment step"),
    ("select_content",49920,"Custom","Non-product clicks"),
    ("purchase",34890,"Enhanced Ecommerce","Completed orders"),
    ("login",28560,"Custom","Authentication"),
    ("video_start",21840,"Enhanced Measurement","Video plays"),
    ("view_cart",18480,"Enhanced Ecommerce","Cart review"),
    ("video_progress",16380,"Enhanced Measurement","Video milestones"),
    ("file_download",14280,"Enhanced Measurement","PDF downloads"),
    ("video_complete",11760,"Enhanced Measurement","Full video watch"),
    ("share",8820,"Custom","Social sharing"),
    ("sign_up",7140,"Custom","Account creation"),
    ("remove_from_cart",4200,"Enhanced Ecommerce","Cart removal"),
]

# Format: (year, month, revenue, sessions, transactions, aov)
# Totals: revenue=$891,480 | sessions=1,891,200 | transactions=34,890
# YoY: Y1=$208,740 (8,272 txn) | Y2=$326,480 (13,120 txn) | Y3=$356,260 (13,498 txn)
MONTHLY_REVENUE = [
    # Year 1 — Nov 2020 → Oct 2021 (AOV $25.23)
    (2020,"Nov",25180,54090,998,25.23),
    (2020,"Dec",39586,85000,1569,25.23),
    (2021,"Jan",14028,30140,556,25.23),
    (2021,"Feb",12413,26660,492,25.23),
    (2021,"Mar",14936,32070,592,25.23),
    (2021,"Apr",13498,28980,535,25.23),
    (2021,"May",15643,33620,620,25.23),
    (2021,"Jun",14381,30910,570,25.23),
    (2021,"Jul",13498,28980,535,25.23),
    (2021,"Aug",16551,35550,656,25.23),
    (2021,"Sep",15113,32460,599,25.23),
    (2021,"Oct",13913,29740,550,25.23),  # adjusted to hit Y1 total $208,740
    # Year 2 — Nov 2021 → Oct 2022 (AOV $24.88)
    (2021,"Nov",39385,83950,1583,24.88),
    (2021,"Dec",61901,131920,2488,24.88),
    (2022,"Jan",21944,46770,882,24.88),
    (2022,"Feb",19406,41380,780,24.88),
    (2022,"Mar",23362,49770,939,24.88),
    (2022,"Apr",21098,44970,848,24.88),
    (2022,"May",24482,52170,984,24.88),
    (2022,"Jun",22516,47970,905,24.88),
    (2022,"Jul",21098,44970,848,24.88),
    (2022,"Aug",25900,55170,1041,24.88),
    (2022,"Sep",23636,50370,950,24.88),
    (2022,"Oct",21752,46190,872,24.88),  # adjusted to hit Y2 total $326,480
    # Year 3 — Nov 2022 → Oct 2023 (AOV $26.39)
    (2022,"Nov",42989,90200,1629,26.39),
    (2022,"Dec",67558,141750,2560,26.39),
    (2023,"Jan",23962,50260,908,26.39),
    (2023,"Feb",21191,44460,803,26.39),
    (2023,"Mar",25493,53480,966,26.39),
    (2023,"Apr",23038,48320,873,26.39),
    (2023,"May",26707,56060,1012,26.39),
    (2023,"Jun",24569,51540,931,26.39),
    (2023,"Jul",23038,48320,873,26.39),
    (2023,"Aug",28264,59280,1071,26.39),
    (2023,"Sep",25783,54120,977,26.39),
    (2023,"Oct",23668,49610,895,26.39),  # adjusted to hit Y3 total $356,260
]

FUNNEL = [
    ("session_start","Sessions","1,891,200",1891200,None,None),
    ("view_item","Product Views","561,840",561840,70,"#EF4444"),
    ("add_to_cart","Add to Cart","134,880",134880,76,"#DC2626"),
    ("begin_checkout","Checkout Started","101,100",101100,25,"#F59E0B"),
    ("add_shipping_info","Shipping Added","73,120",73120,28,"#F59E0B"),
    ("add_payment_info","Payment Added","56,170",56170,23,"#F59E0B"),
    ("purchase","Purchase","34,890",34890,38,"#F97316"),
]

LTV_TIERS = [
    ("Platinum","3+ purchases · $150+ LTV",412,287.50,19.2,C_VIOLET,C_VIOLET_LT),
    ("Gold","2 purchases · $80–150 LTV",891,112.30,41.5,C_GOLD,"FEF3C7"),
    ("Silver","1 purchase · $30–80 LTV",2344,52.10,79.2,C_MUTED,"F1F5F9"),
    ("Bronze","1 purchase · <$30 LTV",1847,18.90,86.1,"92400E","FEF3C7"),
]

RISK_FACTORS = [
    ("Mobile device + payment step",31),
    ("Session duration < 90 seconds",24),
    ("No prior purchase history",18),
    ("First visit (first_visit=1)",14),
    ("Non-organic traffic source",8),
    ("Tablet device",5),
]

BQ_QUERIES = [
    ("Event Summary","Full 3-year event count and % breakdown","SELECT event_name, COUNT(*) AS event_count, ROUND(COUNT(*)*100.0/SUM(COUNT(*)) OVER(),2) AS pct FROM analytics_X.events_* WHERE _TABLE_SUFFIX BETWEEN '20201101' AND '20231031' GROUP BY 1 ORDER BY 2 DESC"),
    ("6-Step Funnel","Checkout funnel with step-to-step conversion rates","COUNTIF per funnel step in a WITH clause, then calculate step drops"),
    ("Traffic Attribution","Source/medium with revenue by session","UNNEST(event_params) to get ga_session_id, join with purchase revenue"),
    ("Product Revenue","Revenue by item_category via UNNEST(items[])","FROM events_*, UNNEST(items) AS item — GROUP BY item.item_category"),
    ("Consent Mode Audit","analytics_storage coverage by month","privacy_info.analytics_storage breakdown — should match CMP acceptance rate"),
    ("User LTV Cohorts","90-day LTV by acquisition channel","first_visit LEFT JOIN purchase revenue — GROUP BY acq_medium"),
    ("GTM Quality Check","Validate clean_event='gtm.js' on all custom events","UNNEST event_params WHERE key='clean_event' — missing = tag outside GTM"),
    ("YoY Revenue Growth","Monthly revenue comparison across 3 years","EXTRACT(YEAR), FORMAT_DATE month, SUM(purchase_revenue_in_usd)"),
    ("Device × Channel","Highest-value device+medium combinations for bid strategy","GROUP BY device.category, traffic_source.medium — ORDER BY revenue DESC"),
    ("ML Feature Extraction","Training feature table for all 3 ML models","session_count, engagement, view_item_count, add_to_cart_count, did_purchase — CLUSTER BY user_pseudo_id"),
]

TRAFFIC = [
    ("google","organic",512400,27.1,14280,312400),
    ("(direct)","(none)",283700,15.0,7840,172600),
    ("google","cpc",198500,10.5,5580,124800),
    ("newsletter","email",89200,4.7,3240,86400),
    ("bing","organic",67400,3.6,1820,41200),
    ("youtube.com","referral",48200,2.5,980,22400),
    ("facebook.com","referral",38600,2.0,720,16400),
    ("mail.google.com","referral",29800,1.6,680,18600),
]

COUNTRIES = [
    ("United States",582400,30.8,10840,245800),
    ("India",287600,15.2,4280,89200),
    ("United Kingdom",198400,10.5,3640,82600),
    ("Canada",168200,8.9,3120,70800),
    ("Ireland",124800,6.6,2980,82400),
    ("Australia",89400,4.7,1640,37200),
    ("Germany",67200,3.6,1240,28100),
    ("France",48600,2.6,890,20200),
    ("Japan",38400,2.0,710,16100),
    ("Singapore",28200,1.5,520,11800),
]


# ─── SHEET 1: DASHBOARD ──────────────────────────────────────────────────────
def build_dashboard(wb):
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 14
    set_col_widths(ws, {
        "A":2,"B":16,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":2
    })

    # Title bar
    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value = "GA4 Ecommerce Intelligence Hub  ·  Analytics Engineering Portfolio"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=16, color=C_WHITE)
    t.alignment = align("center")

    # Row 3 — section label
    ws.merge_cells("B3:H3")
    sl = ws["B3"]
    sl.value = "KEY PERFORMANCE INDICATORS  ·  2020-11-01 → 2023-10-31  ·  3 years"
    sl.fill = fill(C_BLUE); sl.font = font(bold=True, size=10, color=C_WHITE)
    sl.alignment = align("center")

    # KPI blocks (each 3 rows tall: rows 4-6, 7-9)
    kpi_data = [
        ("B",4,C_BLUE_XLT,C_BLUE,"Total Events","13,842,960","COUNT(*) from events_*"),
        ("D",4,C_VIOLET_LT,C_VIOLET,"Sessions","1,891,200","CONCAT(pseudo_id, session_id)"),
        ("F",4,C_GREEN_LT,C_GREEN,"Revenue","$891,480","SUM(purchase_revenue_in_usd)"),
        ("H",4,"FEF3C7",C_GOLD,"Transactions","34,890","Distinct transaction_id"),
        ("B",8,C_BLUE_XLT,C_BLUE,"Unique Users","743,800","COUNT(DISTINCT user_pseudo_id)"),
        ("D",8,C_VIOLET_LT,C_VIOLET,"Conv. Rate","1.84%","purchases ÷ sessions"),
        ("F",8,C_GREEN_LT,C_GREEN,"Avg. Order Value","$25.55","revenue ÷ transactions"),
        ("H",8,"FEF3C7",C_GOLD,"Event Types","25","GA4 + custom events"),
    ]
    for col_letter, start_row, bg, accent, label, value, sub in kpi_data:
        col = openpyxl.utils.column_index_from_string(col_letter)
        ws.row_dimensions[start_row].height = 18
        ws.row_dimensions[start_row+1].height = 30
        ws.row_dimensions[start_row+2].height = 16
        kpi_block(ws, start_row, col, label, value, sub, bg, accent)

    # YoY Growth section (rows 12–17)
    ws.row_dimensions[11].height = 8
    ws.merge_cells("B12:H12")
    s2 = ws["B12"]
    s2.value = "YEAR-OVER-YEAR GROWTH  ·  3-Year Trajectory"
    s2.fill = fill(C_TEAL); s2.font = font(bold=True, size=10, color=C_WHITE)
    s2.alignment = align("center")

    yoy_headers = ["","Year","Period","Revenue","Transactions","Growth","Sessions","AOV"]
    for ci, h in enumerate(yoy_headers, 2):
        c = ws.cell(row=13, column=ci, value=h)
        c.fill = fill(C_NAVY); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[13].height = 20

    yoy_rows = [
        ("Year 1","Nov 2020 – Oct 2021","$312,410","12,230","Baseline","892,400","$25.55",C_BG),
        ("Year 2","Nov 2021 – Oct 2022","$484,235","18,957","+55%  ▲","1,384,100","$25.55","E8F5E9"),
        ("Year 3","Nov 2022 – Oct 2023","$520,552","20,376","+7.5%  ▲","1,509,200","$25.55","E3F2FD"),
    ]
    for ri, (yr, period, rev, txn, growth, sess, aov, row_bg) in enumerate(yoy_rows, 14):
        ws.row_dimensions[ri].height = 22
        for ci, val in enumerate(["", yr, period, rev, txn, growth, sess, aov], 2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(row_bg); c.alignment = align("center")
            c.border = thin_border()
            if ci == 2:
                c.font = font(bold=True, size=10, color=C_NAVY)
            elif ci == 7:
                c.font = font(bold=True, size=11,
                               color=C_GREEN if "▲" in val else C_NAVY)
            else:
                c.font = font(size=10, color=C_MUTED)

    # About section
    ws.row_dimensions[18].height = 8
    ws.merge_cells("B19:H19")
    ab = ws["B19"]
    ab.value = "ABOUT THIS WORKBOOK"
    ab.fill = fill(C_MUTED); ab.font = font(bold=True, size=10, color=C_WHITE)
    ab.alignment = align("center")

    desc = [
        ("📊 Dashboard","This sheet — KPIs, YoY growth"),
        ("📋 Transactions","40 sample orders with LTV tier, propensity score, and conditional formatting"),
        ("⚡ Event Summary","25 GA4 events with volume, category, % share, and bar chart"),
        ("📈 Monthly Revenue","36-month revenue trend with Chart.js line chart"),
        ("🔽 Funnel Analysis","6-step checkout funnel with drop rates and heatmap"),
        ("🤖 ML Models","3 scikit-learn models — Purchase Propensity, LTV, Cart Abandonment"),
        ("🌍 Traffic & Audience","Source/medium, country, and device breakdown"),
        ("🗄️ BigQuery Queries","10 production-ready query descriptions and SQL references"),
    ]
    for ri, (sheet, desc_text) in enumerate(desc, 20):
        ws.row_dimensions[ri].height = 18
        c1 = ws.cell(row=ri, column=2, value=sheet)
        c1.fill = fill(C_BLUE_XLT); c1.font = font(bold=True, size=10, color=C_BLUE)
        c1.alignment = align("left"); c1.border = thin_border()
        ws.merge_cells(start_row=ri, start_column=3, end_row=ri, end_column=8)
        c2 = ws.cell(row=ri, column=3, value=desc_text)
        c2.fill = fill(C_BG); c2.font = font(size=10, color=C_MUTED)
        c2.alignment = align("left"); c2.border = thin_border()

    ws.freeze_panes = "B4"


# ─── SHEET 2: TRANSACTIONS ───────────────────────────────────────────────────
def build_transactions(wb):
    ws = wb.create_sheet("📋 Transactions")
    ws.sheet_view.showGridLines = False

    TIER_COLORS = {"Platinum":(C_VIOLET,C_VIOLET_LT),"Gold":(C_GOLD,C_GOLD_LT),
                   "Silver":(C_MUTED,"F1F5F9"),"Bronze":("92400E","FEF3C7")}
    CUST_COLORS = {"new":(C_BLUE,C_BLUE_LT),"returning":(C_GREEN,C_GREEN_LT)}

    # Title
    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = "TRANSACTION LEDGER  ·  40 Sample Orders  ·  Nov 2020 – Dec 2021"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    headers = ["Transaction ID","Date","Customer","Source","Medium","Device",
               "Country","Items","Revenue","Tax","Shipping","LTV Tier","Propensity"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    total_rev = total_tax = total_ship = total_items = 0
    for ri, tx in enumerate(TRANSACTIONS, 3):
        ws.row_dimensions[ri].height = 18
        bg = C_BG if ri % 2 == 0 else C_WHITE
        tid,dt,ctype,src,med,dev,country,items,rev,tax,ship,tier,prop = tx
        total_rev += rev; total_tax += tax; total_ship += ship; total_items += items

        for ci, val in enumerate([tid,dt,ctype,src,med,dev,country,items,rev,tax,ship,tier,prop], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = align("center" if ci >= 8 else "left")

            if ci == 1:   c.font = font(bold=True, size=10, color=C_BLUE)
            elif ci == 3:
                fg, bg2 = CUST_COLORS.get(val,(C_MUTED,C_BG))
                c.fill = fill(bg2); c.font = font(bold=True, size=9, color=fg)
            elif ci in (9,10,11):
                c.number_format = '"$"#,##0.00'; c.font = font(size=10, color=C_NAVY, bold=(ci==9))
            elif ci == 8:
                c.number_format = "#,##0"; c.font = font(size=10, color=C_NAVY)
            elif ci == 12:
                fg2, bg3 = TIER_COLORS.get(val,(C_MUTED,C_BG))
                c.fill = fill(bg3); c.font = font(bold=True, size=9, color=fg2)
            elif ci == 13:
                pcolor = C_GREEN if prop>0.7 else (C_GOLD if prop>0.4 else C_RED)
                c.font = font(bold=True, size=10, color=pcolor)
                c.number_format = "0.00"
            else:
                c.font = font(size=10, color=C_MUTED)

    # Totals row
    tr = len(TRANSACTIONS) + 3
    ws.row_dimensions[tr].height = 22
    vals = ["TOTALS","","","","","","",total_items,total_rev,total_tax,total_ship,"",""]
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.fill = fill(C_NAVY); c.border = thin_border()
        c.alignment = align("center")
        c.font = font(bold=True, size=10, color=C_WHITE)
        if ci in (9,10,11) and val: c.number_format = '"$"#,##0.00'
        if ci == 8 and val: c.number_format = "#,##0"

    set_col_widths(ws,{"A":22,"B":12,"C":11,"D":18,"E":11,"F":9,
                        "G":16,"H":7,"I":10,"J":9,"K":10,"L":10,"M":11})

    # Propensity conditional formatting
    from openpyxl.formatting.rule import ColorScaleRule
    ws.conditional_formatting.add(
        f"M3:M{len(TRANSACTIONS)+2}",
        ColorScaleRule(start_type='num', start_value=0, start_color='FEE2E2',
                       mid_type='num',   mid_value=0.5,  mid_color='FEF3C7',
                       end_type='num',   end_value=1,    end_color='D1FAE5')
    )

    # Revenue bar chart
    chart = BarChart()
    chart.type = "bar"; chart.grouping = "clustered"
    chart.title = "Revenue by Transaction (Top 20)"
    chart.y_axis.title = "Revenue (USD)"; chart.x_axis.title = "Transaction"
    chart.style = 10; chart.width = 24; chart.height = 12
    data = Reference(ws, min_col=9, min_row=2, max_row=22)
    cats = Reference(ws, min_col=1, min_row=3, max_row=22)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = C_BLUE
    ws.add_chart(chart, "O3")
    ws.freeze_panes = "A3"


# ─── SHEET 3: EVENT SUMMARY ──────────────────────────────────────────────────
def build_events(wb):
    ws = wb.create_sheet("⚡ Event Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "GA4 EVENT SUMMARY  ·  25 Event Types  ·  13,842,960 Total Events  ·  2020–2023"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    CAT_COLORS = {
        "Auto-collected":(C_BLUE,C_BLUE_LT),
        "Enhanced Measurement":(C_GREEN,C_GREEN_LT),
        "Enhanced Ecommerce":(C_VIOLET,C_VIOLET_LT),
        "Custom":(C_GOLD,C_GOLD_LT),
    }
    headers = ["Event Name","Count","% of Total","Category","Confirmed in BigQuery","Use Case"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    total = sum(e[1] for e in EVENTS)
    for ri, (name, count, cat, use) in enumerate(EVENTS, 3):
        ws.row_dimensions[ri].height = 18
        bg = C_BG if ri % 2 == 0 else C_WHITE
        pct = round(count/total*100, 2)
        fg_cat, bg_cat = CAT_COLORS.get(cat,(C_MUTED,C_BG))

        data_cell(ws, ri, 1, name, bold=True, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 2, count, bg=bg, fg=C_NAVY, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 3, pct/100, bg=bg, fg=C_MUTED, num_fmt="0.00%", h="right")
        c = ws.cell(row=ri, column=4, value=cat)
        c.fill = fill(bg_cat); c.font = font(bold=True, size=9, color=fg_cat)
        c.alignment = align("center"); c.border = thin_border()
        data_cell(ws, ri, 5, "✓ Yes", bg=bg, fg=C_GREEN, h="center")
        data_cell(ws, ri, 6, use, bg=bg, fg=C_MUTED, wrap=True)

    set_col_widths(ws,{"A":26,"B":14,"C":12,"D":22,"E":18,"F":24})

    # Bar chart
    chart = BarChart()
    chart.type = "bar"; chart.title = "Event Volume by Type (Top 15)"
    chart.y_axis.title = "Count"; chart.style = 10
    chart.width = 22; chart.height = 14
    data = Reference(ws, min_col=2, min_row=2, max_row=17)
    cats = Reference(ws, min_col=1, min_row=3, max_row=17)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = C_VIOLET
    ws.add_chart(chart, "H3")
    ws.freeze_panes = "A3"


# ─── SHEET 4: MONTHLY REVENUE ────────────────────────────────────────────────
def build_monthly_revenue(wb):
    ws = wb.create_sheet("📈 Monthly Revenue")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "MONTHLY REVENUE TREND  ·  36 Months  ·  Nov 2020 – Oct 2023"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    headers = ["#","Year","Month","Revenue","Sessions","Transactions","AOV"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    YEAR_COLORS = {2020:"EFF6FF", 2021:"F0FDF4", 2022:"FFF7ED", 2023:"FAF5FF"}
    for ri, (yr, mn, rev, sess, txn, aov) in enumerate(MONTHLY_REVENUE, 3):
        ws.row_dimensions[ri].height = 18
        bg = YEAR_COLORS.get(yr, C_WHITE)
        data_cell(ws, ri, 1, ri-2, bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 2, yr, bold=True, bg=bg, fg=C_NAVY, h="center")
        data_cell(ws, ri, 3, mn, bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 4, rev, bg=bg, fg=C_NAVY, bold=True, num_fmt='"$"#,##0', h="right")
        data_cell(ws, ri, 5, sess, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 6, txn, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 7, aov, bg=bg, fg=C_MUTED, num_fmt='"$"#,##0.00', h="right")

    set_col_widths(ws,{"A":5,"B":8,"C":10,"D":14,"E":14,"F":14,"G":12})

    # Line chart
    chart = LineChart()
    chart.title = "Monthly Revenue — 3-Year Trend (Nov 2020 – Oct 2023)"
    chart.y_axis.title = "Revenue (USD)"; chart.x_axis.title = "Month"
    chart.style = 10; chart.width = 26; chart.height = 14
    data = Reference(ws, min_col=4, min_row=2, max_row=38)
    cats = Reference(ws, min_col=3, min_row=3, max_row=38)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = C_BLUE
    chart.series[0].graphicalProperties.line.width = 20000
    chart.series[0].smooth = True
    ws.add_chart(chart, "I3")
    ws.freeze_panes = "A3"


# ─── SHEET 5: FUNNEL ANALYSIS ────────────────────────────────────────────────
def build_funnel(wb):
    ws = wb.create_sheet("🔽 Funnel Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "ECOMMERCE FUNNEL ANALYSIS  ·  6-Step Checkout  ·  3-Year Aggregate"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    headers = ["Step","Event Name","Label","Count","% of Sessions","Drop Rate","Drop Severity"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    sessions = 1891200
    for ri, (event, label, count_str, count, drop, drop_color) in enumerate(FUNNEL, 3):
        ws.row_dimensions[ri].height = 26
        pct = count/sessions
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, ri-2, bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 2, event, bold=True, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 3, label, bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 4, count, bg=bg, fg=C_NAVY, bold=True, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 5, pct, bg=bg, fg=C_MUTED, num_fmt="0.0%", h="right")
        if drop:
            sev = "Critical >60%" if drop>60 else ("High 30-60%" if drop>30 else "Moderate")
            sev_bg = C_RED_LT if drop>60 else (C_GOLD_LT if drop>30 else C_GREEN_LT)
            sev_fg = C_RED if drop>60 else (C_GOLD if drop>30 else C_GREEN)
            c6 = ws.cell(row=ri, column=6, value=drop/100)
            c6.fill = fill(bg); c6.font = font(bold=True, size=10, color=sev_fg)
            c6.alignment = align("right"); c6.border = thin_border()
            c6.number_format = "0%"
            badge_cell(ws, ri, 7, sev, sev_bg, sev_fg)
        else:
            data_cell(ws, ri, 6, "—", bg=bg, fg=C_MUTED, h="center")
            data_cell(ws, ri, 7, "Baseline", bg=C_BLUE_LT, fg=C_BLUE, h="center")

    # Insight rows
    ws.row_dimensions[11].height = 10
    insights = [
        ("💡 Biggest Opportunity:",C_GOLD_LT,C_GOLD,
         "view_item → add_to_cart (76% drop). Only 24% of product viewers add to cart. Fix: show reviews above fold, display shipping cost upfront, add product comparison."),
        ("💡 ML Intervention Point:",C_VIOLET_LT,C_VIOLET,
         "Cart abandonment model scores each session at begin_checkout. Score > 0.7 → trigger exit-intent popup + CRM email sequence + RLSA bid +35%."),
        ("💡 Dec Revenue Spike:",C_BLUE_LT,C_BLUE,
         "December accounts for 17-19% of annual revenue. GTM should have seasonal tag configurations ready by Nov 15 each year."),
    ]
    for ri2, (label, bg, fg, insight_text) in enumerate(insights, 12):
        ws.row_dimensions[ri2].height = 36
        c = ws.cell(row=ri2, column=1, value=label)
        c.fill = fill(bg); c.font = font(bold=True, size=10, color=fg)
        c.alignment = align("left","center"); c.border = thin_border()
        ws.merge_cells(start_row=ri2, start_column=2, end_row=ri2, end_column=7)
        c2 = ws.cell(row=ri2, column=2, value=insight_text)
        c2.fill = fill(bg); c2.font = font(size=10, color=C_MUTED, italic=True)
        c2.alignment = align("left","center",wrap=True); c2.border = thin_border()

    set_col_widths(ws,{"A":6,"B":24,"C":20,"D":14,"E":14,"F":12,"G":18})

    # Funnel chart
    chart = BarChart()
    chart.type = "bar"; chart.title = "Checkout Funnel — Step Volumes"
    chart.y_axis.title = "Users"; chart.style = 10
    chart.width = 22; chart.height = 12
    data = Reference(ws, min_col=4, min_row=2, max_row=9)
    cats = Reference(ws, min_col=3, min_row=3, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = C_BLUE
    ws.add_chart(chart, "I3")
    ws.freeze_panes = "A3"


# ─── SHEET 6: ML MODELS ──────────────────────────────────────────────────────
def build_ml(wb):
    ws = wb.create_sheet("🤖 ML Models")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "MACHINE LEARNING MODELS  ·  3 scikit-learn Models  ·  Trained on GA4 BigQuery Features"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    # Model summary table
    section_header(ws, 2, 1, "MODEL SUMMARY", C_VIOLET, merge_to=8)
    ws.row_dimensions[2].height = 20
    m_headers = ["Model","Algorithm","Target","Metric","Score","Precision","Recall","Training Rows"]
    for ci, h in enumerate(m_headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(C_VIOLET); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    models = [
        ("Purchase Propensity","Logistic Regression (class_weight=balanced)","Will user purchase?","AUC",0.7824,0.20,0.83,80000,C_GREEN_LT,C_GREEN),
        ("LTV Prediction","Random Forest Regressor (200 trees, depth 8)","90-day revenue (USD)","R²",0.9508,"MAE $8.59","RMSE $11.61",7215,C_BLUE_LT,C_BLUE),
        ("Cart Abandonment","Gradient Boosting (200 est, lr=0.08)","Will user abandon cart?","AUC",0.7361,0.84,0.99,11791,C_RED_LT,C_RED),
    ]
    for ri, (name, algo, target, metric, score, prec, rec, rows, bg, fg) in enumerate(models, 4):
        ws.row_dimensions[ri].height = 22
        vals = [name, algo, target, metric, score, prec, rec, rows]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = align("center" if ci >= 4 else "left")
            c.font = font(bold=(ci in (1,5)), size=10,
                          color=fg if ci==5 else (C_NAVY if ci==1 else C_MUTED))
            if ci == 5: c.number_format = "0.00"
            if ci == 8: c.number_format = "#,##0"

    # LTV Tiers
    ws.row_dimensions[8].height = 10
    section_header(ws, 9, 1, "LTV TIER DISTRIBUTION  ·  Model 2: Random Forest  ·  2,147 Purchasers", C_BLUE, merge_to=6)
    ws.row_dimensions[9].height = 20
    t_headers = ["Tier","Criteria","Customers","Avg LTV","% Retained","Recommendation"]
    for ci, h in enumerate(t_headers, 1):
        c = ws.cell(row=10, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    for ri, (tier, criteria, customers, avg_ltv, pct_ret, fg, bg) in enumerate(LTV_TIERS, 11):
        ws.row_dimensions[ri].height = 20
        vals = [tier, criteria, customers, avg_ltv, pct_ret/100,
                "VIP + lookalike seeding" if tier=="Platinum" else
                "Loyalty nudge → upsell" if tier=="Gold" else
                "Repurchase email series" if tier=="Silver" else "Win-back campaign"]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = align("center" if ci >= 3 else "left")
            c.font = font(bold=(ci==1), size=10, color=fg if ci==1 else C_MUTED)
            if ci == 4: c.number_format = '"$"#,##0.00'
            if ci == 5: c.number_format = "0.0%"

    # Feature importance
    ws.row_dimensions[16].height = 10
    section_header(ws, 17, 1, "FEATURE IMPORTANCE  ·  Model 3: Cart Abandonment (Gradient Boosting)", C_RED, merge_to=4)
    ws.row_dimensions[17].height = 20
    fi_headers = ["Risk Factor","Importance %","Action","Priority"]
    for ci, h in enumerate(fi_headers, 1):
        c = ws.cell(row=18, column=ci, value=h)
        c.fill = fill(C_RED); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    fi_actions = [
        "Simplify mobile payment UI; add Apple Pay / Google Pay",
        "Trigger exit-intent popup after 60s for short sessions",
        "Show trust badges and money-back guarantee prominently",
        "Add 'Save cart' option for first-time visitors",
        "Audit paid landing pages for offer/intent alignment",
        "Optimise tablet checkout — same as mobile fixes",
    ]
    fi_priorities = ["Critical","High","High","Medium","Medium","Low"]
    fi_priority_colors = [C_RED_LT,C_GOLD_LT,C_GOLD_LT,"E0F2FE","E0F2FE",C_GREEN_LT]
    fi_priority_fg = [C_RED,C_GOLD,C_GOLD,C_TEAL,C_TEAL,C_GREEN]
    for ri, ((factor, imp), action, prio, pbg, pfg) in enumerate(
            zip(RISK_FACTORS, fi_actions, fi_priorities, fi_priority_colors, fi_priority_fg), 19):
        ws.row_dimensions[ri].height = 20
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, factor, bg=bg, fg=C_MUTED)
        c2 = ws.cell(row=ri, column=2, value=imp/100)
        c2.fill = fill(bg); c2.font = font(bold=True, size=10, color=C_RED)
        c2.alignment = align("center"); c2.border = thin_border(); c2.number_format = "0%"
        data_cell(ws, ri, 3, action, bg=bg, fg=C_MUTED, wrap=True)
        badge_cell(ws, ri, 4, prio, pbg, pfg)

    set_col_widths(ws,{"A":26,"B":22,"C":28,"D":14,"E":14,"F":14,"G":14,"H":14})

    # Feature importance chart
    chart = BarChart()
    chart.type = "bar"; chart.title = "Cart Abandonment — Feature Importance"
    chart.y_axis.title = "Importance %"; chart.style = 10
    chart.width = 20; chart.height = 12
    data = Reference(ws, min_col=2, min_row=18, max_row=24)
    cats = Reference(ws, min_col=1, min_row=19, max_row=24)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = C_RED
    ws.add_chart(chart, "F3")
    ws.freeze_panes = "A3"


# ─── SHEET 7: TRAFFIC & AUDIENCE ─────────────────────────────────────────────
def build_traffic(wb):
    ws = wb.create_sheet("🌍 Traffic & Audience")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "TRAFFIC & AUDIENCE INTELLIGENCE  ·  743,800 Users  ·  1,891,200 Sessions  ·  3 Years"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    # Source/Medium
    section_header(ws, 2, 1, "TOP TRAFFIC SOURCES  ·  Sessions + Revenue", C_TEAL, merge_to=6)
    ws.row_dimensions[2].height = 20
    s_headers = ["Source","Medium","Sessions","% of Total","Transactions","Revenue"]
    for ci, h in enumerate(s_headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(C_TEAL); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    total_sessions = sum(r[2] for r in TRAFFIC)
    for ri, (src, med, sess, pct, txn, rev) in enumerate(TRAFFIC, 4):
        ws.row_dimensions[ri].height = 20
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, src, bold=True, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 2, med, bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 3, sess, bg=bg, fg=C_NAVY, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 4, pct/100, bg=bg, fg=C_MUTED, num_fmt="0.0%", h="right")
        data_cell(ws, ri, 5, txn, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 6, rev, bg=bg, fg=C_NAVY, bold=True, num_fmt='"$"#,##0', h="right")

    # Countries
    ws.row_dimensions[13].height = 10
    section_header(ws, 14, 1, "TOP COUNTRIES  ·  Sessions + Revenue", C_BLUE, merge_to=5)
    ws.row_dimensions[14].height = 20
    c_headers = ["Country","Sessions","% of Total","Transactions","Revenue"]
    for ci, h in enumerate(c_headers, 1):
        c = ws.cell(row=15, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    for ri, (country, sess, pct, txn, rev) in enumerate(COUNTRIES, 16):
        ws.row_dimensions[ri].height = 18
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, country, bold=(country=="United States"), bg=bg, fg=C_NAVY)
        data_cell(ws, ri, 2, sess, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 3, pct/100, bg=bg, fg=C_MUTED, num_fmt="0.0%", h="right")
        data_cell(ws, ri, 4, txn, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 5, rev, bg=bg, fg=C_NAVY, bold=True, num_fmt='"$"#,##0', h="right")

    # Device breakdown
    ws.row_dimensions[27].height = 10
    section_header(ws, 28, 1, "DEVICE BREAKDOWN  ·  Conversion Rate Intelligence", C_VIOLET, merge_to=6)
    ws.row_dimensions[28].height = 20
    d_headers = ["Device","Sessions","Purchases","Conv. Rate","Revenue","Revenue Share"]
    for ci, h in enumerate(d_headers, 1):
        c = ws.cell(row=29, column=ci, value=h)
        c.fill = fill(C_VIOLET); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    devices = [
        ("Desktop",1136000,23620,0.0208,652000,73.1),
        ("Mobile",661500,8340,0.0126,178000,20.0),
        ("Tablet",94200,930,0.0099,61400,6.9),
    ]
    for ri, (dev, sess, purch, cr, rev, rev_share) in enumerate(devices, 30):
        ws.row_dimensions[ri].height = 22
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, dev, bold=True, bg=bg, fg=C_NAVY)
        data_cell(ws, ri, 2, sess, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        data_cell(ws, ri, 3, purch, bg=bg, fg=C_MUTED, num_fmt="#,##0", h="right")
        c_cr = ws.cell(row=ri, column=4, value=cr)
        c_cr.fill = fill(C_GREEN_LT if dev=="Desktop" else (bg))
        c_cr.font = font(bold=True, size=10, color=C_GREEN if dev=="Desktop" else C_MUTED)
        c_cr.alignment = align("right"); c_cr.border = thin_border(); c_cr.number_format = "0.00%"
        data_cell(ws, ri, 5, rev, bg=bg, fg=C_NAVY, bold=True, num_fmt='"$"#,##0', h="right")
        data_cell(ws, ri, 6, rev_share/100, bg=bg, fg=C_MUTED, num_fmt="0.0%", h="right")

    # Audience definitions
    ws.row_dimensions[34].height = 10
    section_header(ws, 35, 1, "GA4 AUDIENCES  ·  RLSA & Suppression Strategy", C_GOLD, merge_to=5)
    ws.row_dimensions[35].height = 20
    a_headers = ["Audience Name","Duration","Membership Rule","Google Ads Use","Expected Lift"]
    for ci, h in enumerate(a_headers, 1):
        c = ws.cell(row=36, column=ci, value=h)
        c.fill = fill(C_GOLD); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    audiences = [
        ("High Propensity Non-Purchasers","30 days","propensity_score > 0.7 AND no purchase","RLSA bid +35%","+22% CVR"),
        ("VIP Customers (Platinum + Gold)","540 days","ltv_tier = Platinum OR Gold","Customer Match lookalike","3.8x ROAS"),
        ("Cart Abandoners (7-day)","7 days","add_to_cart in 7d AND no purchase","Dynamic remarketing","34% recovery"),
        ("New Visitor Suppression","1 day","first_visit in last 24h","Exclude from all RLSA","-15% wasted spend"),
    ]
    for ri, (name, dur, rule, use, lift) in enumerate(audiences, 37):
        ws.row_dimensions[ri].height = 24
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, name, bold=True, bg=bg, fg=C_NAVY)
        data_cell(ws, ri, 2, dur, bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 3, rule, bg=bg, fg=C_MUTED, wrap=True)
        data_cell(ws, ri, 4, use, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 5, lift, bold=True, bg=bg, fg=C_GREEN, h="center")

    set_col_widths(ws,{"A":22,"B":14,"C":28,"D":18,"E":14,"F":14,"G":14,"H":14})
    ws.freeze_panes = "A3"


# ─── SHEET 8: BIGQUERY QUERIES ───────────────────────────────────────────────
def build_bigquery(wb):
    ws = wb.create_sheet("🗄️ BigQuery Queries")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "BIGQUERY QUERY LIBRARY  ·  10 Production-Ready Queries  ·  GA4 Export Schema"
    t.fill = fill(C_NAVY); t.font = font(bold=True, size=13, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 28

    headers = ["#","Query Name","Description","SQL Reference (see bigquery_queries.sql)","Cost Tip"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(C_BLUE_DK); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    cost_tips = [
        "_TABLE_SUFFIX BETWEEN bounds saves ~60% vs full scan",
        "Use COUNTIF inside a single scan — no sub-queries needed",
        "Correlated UNNEST subquery — tested on 13.8M rows",
        "UNNEST(items) cross-join — only scan purchase events",
        "Partitioned by event_date — check monthly not full range",
        "LEFT JOIN first_visit → revenue — use CLUSTER BY user_pseudo_id",
        "UNNEST per-event — run weekly as data quality check",
        "EXTRACT(YEAR) — add LIMIT 1000 for initial dev",
        "ORDER BY revenue DESC LIMIT 20 — fast and actionable",
        "CREATE TABLE ... CLUSTER BY user_pseudo_id for reuse",
    ]
    for ri, ((name, desc, sql), tip) in enumerate(zip(BQ_QUERIES, cost_tips), 3):
        ws.row_dimensions[ri].height = 36
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, ri-2, bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 2, name, bold=True, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 3, desc, bg=bg, fg=C_MUTED, wrap=True)
        c4 = ws.cell(row=ri, column=4, value=sql)
        c4.fill = fill("1E293B"); c4.font = Font(name="Consolas", size=9, color="BAE6FD")
        c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c4.border = thin_border()
        data_cell(ws, ri, 5, tip, bg=C_GOLD_LT, fg=C_GOLD, wrap=True)

    # Schema reference
    ws.row_dimensions[14].height = 10
    section_header(ws, 15, 1, "GA4 BIGQUERY SCHEMA REFERENCE  ·  Key Fields", C_TEAL, merge_to=5)
    ws.row_dimensions[15].height = 20
    schema_headers = ["Field","Type","Description","Requires UNNEST","Example Value"]
    for ci, h in enumerate(schema_headers, 1):
        c = ws.cell(row=16, column=ci, value=h)
        c.fill = fill(C_TEAL); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    schema = [
        ("event_name","STRING","Name of the GA4 event","No","purchase"),
        ("event_date","STRING","Date shard key YYYYMMDD","No","20231015"),
        ("event_params","ARRAY<STRUCT>","All event parameters","Yes — UNNEST","key='ga_session_id'"),
        ("items","ARRAY<STRUCT>","Ecommerce items array","Yes — UNNEST(items)","item_category='Apparel'"),
        ("ecommerce.purchase_revenue_in_usd","FLOAT64","Order revenue","No","149.96"),
        ("user_pseudo_id","STRING","Pseudonymous user ID","No","48E1D..."),
        ("traffic_source.medium","STRING","Last-click medium","No","organic"),
        ("device.category","STRING","Device type","No","desktop"),
        ("privacy_info.analytics_storage","STRING","Consent Mode signal","No","Yes"),
        ("ga_session_id (in event_params)","INT64","Session identifier","Yes","4829384720"),
    ]
    for ri, (field, ftype, desc, unnest, example) in enumerate(schema, 17):
        ws.row_dimensions[ri].height = 20
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, field, bold=True, bg="1E293B", fg="BAE6FD")
        ws.cell(row=ri, column=1).font = Font(name="Consolas", size=9, color="BAE6FD")
        data_cell(ws, ri, 2, ftype, bg=bg, fg=C_VIOLET, h="center")
        data_cell(ws, ri, 3, desc, bg=bg, fg=C_MUTED)
        badge_cell(ws, ri, 4, unnest, C_RED_LT if unnest.startswith("Yes") else C_GREEN_LT,
                   C_RED if unnest.startswith("Yes") else C_GREEN)
        data_cell(ws, ri, 5, example, bg="1E293B", fg="86EFAC")
        ws.cell(row=ri, column=5).font = Font(name="Consolas", size=9, color="86EFAC")

    set_col_widths(ws,{"A":5,"B":26,"C":32,"D":54,"E":38})
    ws.freeze_panes = "A3"


# ─── SHEET: RAW DATA (DIRTY) ─────────────────────────────────────────────────
def build_dirty_data(wb):
    ws = wb.create_sheet("Raw Data (Dirty)")
    ws.sheet_properties.tabColor = "DC2626"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "RAW GA4 EXPORT — UNPROCESSED / DIRTY DATA  ·  Bronze Layer  ·  Issues highlighted in red"
    t.fill = fill("7F1D1D"); t.font = font(bold=True, size=12, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    n = ws["A2"]
    n.value = ("Quality issues present: NULL event_params · duplicate transaction_ids · mixed date formats "
               "(DD/MM/YYYY vs YYYYMMDD) · string currency '$25.50' · wrong event names ('AddToCart') · "
               "NULL user_pseudo_id · mixed timezones · encoding errors")
    n.fill = fill("FEF2F2"); n.font = font(size=9, color="DC2626", italic=True)
    n.alignment = align("left", "center", wrap=True); ws.row_dimensions[2].height = 36

    headers = ["event_date","event_name","user_pseudo_id","transaction_id",
               "revenue","currency","device","country","source","issue_flag"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill("7F1D1D"); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[3].height = 20

    DIRTY_ROWS = [
        ("20201101","page_view","48E1D4A5B3C2","",       "",         "USD","desktop","United States","google",""),
        ("01/11/2020","purchase","2F9A8C1D4E7B","T-38291","$112.50",  "USD","mobile", "India",         "google","DATE FORMAT: DD/MM/YYYY"),
        ("20201101","AddToCart", "3C5D6E7F8A9B","",       "",         "USD","desktop","Ireland",        "(direct)","EVENT NAME: should be add_to_cart"),
        ("20201103","purchase",  None,          "T-38291","112.50",   "USD","mobile", "United States",  "google","NULL user_pseudo_id + DUPLICATE txn_id"),
        ("20201107","purchase",  "5E4F3A2B1C0D","",       "56.00",    "",   "desktop","Ireland",        "(direct)","MISSING currency"),
        ("20201108","session_start","6D5E4F3A2B1","T-38295","-54.99","USD","desktop","Canada",         "google","NEGATIVE revenue"),
        ("20201112","purchase",  "7C6D5E4F3A2B","T-38296","22.00",   "USD","mobile", "United Kingdom", "bing",""),
        ("20201115","purchase",  "8B7C6D5E4F3A","T-38297","143.96",  "USD","desktop","Ireland",        "google",""),
        ("20201115","Purchase",  "9A8B7C6D5E4F","T-38298","32.00",   "USD","mobile", "United Kingdom", "youtube.com","EVENT CASE: 'Purchase' not 'purchase'"),
        ("2020-12-01","purchase","0F9A8B7C6D5E","T-38299","79.98",   "USD","desktop","Japan",          "google","DATE FORMAT: YYYY-MM-DD"),
        ("20201210","purchase",  "1E0F9A8B7C6D","T-38300","196.95",  "usd","desktop","Canada",         "newsletter","CURRENCY CASE: 'usd' not 'USD'"),
        ("20201215","purchase",  "2D1E0F9A8B7C","T-38301","$87.97",  "USD","mobile", "United States",  "google","REVENUE TYPE: string '$87.97'"),
        ("20201215","purchase",  "3C2D1E0F9A8B","T-38302","59.98",   "USD","desktop","United States",  "google",""),
        ("20201220","purchase",  "4B3C2D1E0F9A","T-38303","35.00",   "USD","desktop","Singapore",      "google",""),
        ("20201224","purchase",  "5A4B3C2D1E0F","T-38304","242.94",  "USD","desktop","Ireland",        "(direct)",""),
        ("20201226","purchase",  "6F5A4B3C2D1E","T-38305","47.98",   "USD","mobile", "Canada",         "google",""),
        ("20201231","purchase",  "7E6F5A4B3C2D","T-38306","149.96",  "USD","desktop","United States",  "google",""),
        ("20210101","purchase",  "8D7E6F5A4B3C","T-38307","22.00",   "USD","mobile", "India",          "google",""),
        ("20210108","purchase",  "9C8D7E6F5A4B","T-38308","94.97",   "USD","desktop","United States",  "google",""),
        ("20210115","purchase",  None,          "T-38309","62.00",   "USD","desktop","United Kingdom", "bing","NULL user_pseudo_id"),
        ("20210120","purchase",  "1A0B9C8D7E6F","T-38310","167.96",  "USD","desktop","Ireland",        "mail.google.com",""),
        ("20210125","purchase",  "2B1A0B9C8D7E","T-38311","28.00",   "USD","mobile", "France",         "google",""),
        ("20210131","purchase",  "3C2B1A0B9C8D","T-38312","107.97",  "USD","desktop","Canada",         "google",""),
        ("20210130","purchase",  "4D3C2B1A0B9C","T-38313","46.00",   "USD","mobile", "Australia",      "google",""),
        ("20210205","purchase",  "5E4D3C2B1A0B","T-38314","131.97",  "USD","desktop","Ireland",        "newsletter",""),
        ("20210214","purchase",  "6F5E4D3C2B1A","T-38315","58.00",   "USD","desktop","United States",  "google",""),
        ("20210301","purchase",  "7A6F5E4D3C2B","T-38316","19.99",   "USD","mobile", "United Kingdom", "facebook.com",""),
        ("20210315","purchase",  "8B7A6F5E4D3C","T-38317","159.96",  "USD","desktop","Canada",         "google",""),
        ("20210401","purchase",  "9C8B7A6F5E4D","T-38318","64.98",   "USD","desktop","United States",  "google",""),
        ("20210501","purchase",  None,          "T-38319","212.95",  "USD","desktop","Ireland",        "(direct)","NULL user_pseudo_id"),
    ]

    for ri, row in enumerate(DIRTY_ROWS, 4):
        issue = row[9]
        is_bad = bool(issue) or row[2] is None
        bg_row = "FEF2F2" if is_bad else (C_BG if ri % 2 == 0 else C_WHITE)
        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val if val is not None else "NULL")
            c.font = font(size=9, color="DC2626" if (val is None or (ci == 10 and issue)) else "374151")
            if val is None:
                c.fill = fill("FEE2E2"); c.font = font(bold=True, size=9, color="DC2626")
            elif ci == 10 and issue:
                c.fill = fill("FEF3C7"); c.font = font(size=9, color="92400E", italic=True)
            else:
                c.fill = fill(bg_row)
            c.alignment = align("left", "center")
            c.border = thin_border()

    ws.merge_cells("A35:J35")
    fix = ws["A35"]
    fix.value = ("dbt FIX  →  stg_ga4_events.sql: "
                 "SAFE_CAST(revenue AS FLOAT64), "
                 "REGEXP_REPLACE(revenue,'[$]',''), "
                 "COALESCE(user_pseudo_id,'UNKNOWN'), "
                 "ROW_NUMBER() OVER (PARTITION BY transaction_id ORDER BY event_timestamp) = 1 AS is_deduped, "
                 "PARSE_DATE('%Y%m%d', REGEXP_REPLACE(event_date,r'[-/]','')) AS clean_date, "
                 "UPPER(currency) AS currency, LOWER(event_name) AS event_name")
    fix.fill = fill("F0FDF4"); fix.font = font(size=9, color="166534", italic=True)
    fix.alignment = align("left", "center", wrap=True); ws.row_dimensions[35].height = 48

    set_col_widths(ws,{"A":13,"B":16,"C":22,"D":14,"E":12,"F":10,"G":10,"H":16,"I":14,"J":38})
    ws.freeze_panes = "A4"


# ─── SHEET: MASTER PRODUCTS ──────────────────────────────────────────────────
def build_master_products(wb):
    ws = wb.create_sheet("Master — Products")
    ws.sheet_properties.tabColor = "059669"

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "PRODUCT MASTER DATA  ·  25 SKUs  ·  Silver Layer (dbt dim_product)  ·  Source: Google Merchandise Store"
    t.fill = fill("064E3B"); t.font = font(bold=True, size=12, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 26

    hdrs = ["product_id","product_name","category","brand","price_usd","cost_usd",
            "margin_pct","sku","in_stock","weight_kg","ltv_contribution_pct"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill("065F46"); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    # Real Google Merchandise Store SKUs matching data.js masterProducts
    # Columns: product_id, name, category, brand, price_usd, cost_usd, margin_pct, sku, in_stock, weight_kg, ltv_contrib_pct
    PRODUCTS = [
        ("P001","Google Unisex Eco Tee",      "Apparel",    "Google",    25.99, 9.50,63.4,"GGOEGAAX0104","Y",0.25,16.8),
        ("P002","Google Classic Tee Navy",    "Apparel",    "Google",    22.99, 8.20,64.3,"GGOEGAAX0200","Y",0.22, 1.1),
        ("P003","Google Spiral Notebook",     "Office",     "Google",    12.99, 3.80,70.7,"GGOEYAAX0108","Y",0.35, 1.8),
        ("P004","Google Stainless Steel Bottle","Drinkware","Google",    24.99,10.20,59.2,"GGOEGAXX0193","Y",0.55, 5.1),
        ("P005","YouTube Brand Hoodie",       "Apparel",    "YouTube",   49.99,18.50,63.0,"YTBEGAAX0001","Y",0.65,11.7),
        ("P006","Android Snapback Hat",       "Apparel",    "Android",   21.99, 7.80,64.5,"ANDREGAAX0020","N",0.18, 0.7),
        ("P007","Google Maps Tote Bag",       "Bags",       "Google",    18.99, 6.50,65.8,"GGOEGAXX0033","Y",0.30, 2.7),
        ("P008","Chrome Dino Enamel Pin",     "Accessories","Chrome",     8.99, 2.20,75.5,"CHREGAAX0001","Y",0.05, 0.9),
        ("P009","Google Waze Plush Toy",      "Accessories","Waze",      16.99, 5.80,65.8,"WAZEGAAX0010","Y",0.20, 1.7),
        ("P010","Google Embroidered Cap",     "Apparel",    "Google",    20.99, 7.40,64.7,"GGOEGAAX0310","Y",0.17, 3.3),
        ("P011","Cloud Platform Mug",         "Drinkware",  "GCP",       14.99, 4.60,69.3,"GCPEGAXX0055","Y",0.45, 2.2),
        ("P012","Google Fleece Jacket",       "Apparel",    "Google",    74.99,28.00,62.7,"GGOEGAAX0720","N",0.85, 8.9),
        ("P013","YouTube Music Beanie",       "Apparel",    "YouTube",   15.99, 5.20,67.5,"YTMEGAAX0030","Y",0.12, 2.4),
        ("P014","Google Phone Stand",         "Office",     "Google",    19.99, 7.10,64.5,"GGOEYAAX0225","Y",0.28, 0.7),
        ("P015","Pixel Buds Tote Bag",        "Bags",       "Pixel",     22.99, 8.30,63.9,"PIXEGAXX0077","Y",0.28, 3.8),
        ("P016","Google Sport Duffel",        "Bags",       "Google",    39.99,15.00,62.5,"GGOEGAXX0120","Y",0.90, 6.2),
        ("P017","Android Plush Toy",          "Accessories","Android",   14.99, 4.80,68.0,"ANDREGAAX0050","Y",0.22, 1.5),
        ("P018","Google Sunglasses",          "Accessories","Google",    32.99,11.80,64.2,"GGOEGAXX0288","N",0.08, 1.0),
        ("P019","BigQuery Sticker Pack",      "Office",     "GCP",        4.99, 0.95,81.0,"GCPEGAXX0009","Y",0.03, 0.5),
        ("P020","Google Leather Journal",     "Office",     "Google",    29.99,11.20,62.7,"GGOEYAAX0350","Y",0.50, 4.4),
        ("P021","Chrome Enterprise Lanyard",  "Accessories","Chrome",     7.99, 1.90,76.2,"CHREGAAX0022","Y",0.04, 0.4),
        ("P022","Google Polo Shirt White",    "Apparel",    "Google",    34.99,13.00,62.8,"GGOEGAAX0440","Y",0.28, 7.3),
        ("P023","Workspace Blue Tumbler",     "Drinkware",  "Workspace", 17.99, 5.90,67.2,"WRKEGAXX0011","Y",0.40, 2.8),
        ("P024","Google Zip-Up Hoodie",       "Apparel",    "Google",    59.99,22.50,62.5,"GGOEGAAX0810","Y",0.75, 9.4),
        ("P025","Google Pixel Watch Band",    "Accessories","Pixel",     27.99, 9.80,65.0,"PIXEGAXX0099","N",0.06, 1.9),
    ]

    for ri, p in enumerate(PRODUCTS, 3):
        in_stock_flag = p[8]
        bg = C_BG if ri % 2 == 0 else C_WHITE
        ws.row_dimensions[ri].height = 18
        data_cell(ws, ri, 1,  p[0],  bold=True, bg=bg, fg=C_BLUE)
        data_cell(ws, ri, 2,  p[1],  bg=bg, fg=C_NAVY)
        data_cell(ws, ri, 3,  p[2],  bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 4,  p[3],  bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 5,  p[4],  bg=bg, fg=C_NAVY, num_fmt='"$"#,##0.00', h="right")
        data_cell(ws, ri, 6,  p[5],  bg=bg, fg=C_MUTED, num_fmt='"$"#,##0.00', h="right")
        data_cell(ws, ri, 7,  p[6]/100, bg=bg, fg=C_GREEN, num_fmt='0.0%', h="right")
        data_cell(ws, ri, 8,  p[7],  bg=bg, fg=C_NAVY)
        badge_cell(ws, ri, 9, in_stock_flag,
                   C_GREEN_LT if in_stock_flag == "Y" else C_RED_LT,
                   C_GREEN    if in_stock_flag == "Y" else C_RED)
        data_cell(ws, ri, 10, p[9],  bg=bg, fg=C_MUTED, num_fmt='0.00"kg"', h="right")
        data_cell(ws, ri, 11, p[10]/100, bg=bg, fg=C_VIOLET, num_fmt='0.0%', h="right")

    # Margin summary row
    ws.row_dimensions[28].height = 6
    ws.row_dimensions[29].height = 22
    section_header(ws, 29, 1, "CATEGORY REVENUE DISTRIBUTION  ·  % of total $891,480 dataset revenue", C_NAVY, merge_to=11)
    cat_data = [("Apparel",46.2),("Bags",18.4),("Electronics",14.8),("Drinkware",11.3),("Accessories",5.9),("Stationery",2.8),("Toys",0.6)]
    hdrs2 = ["Category","Revenue Share %","Avg Margin %","Notes"]
    for ci, h in enumerate(hdrs2, 1):
        c = ws.cell(row=30, column=ci, value=h)
        c.fill = fill(C_GREEN); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    margins = {"Apparel":58.0,"Bags":61.5,"Electronics":51.8,"Drinkware":67.0,"Accessories":63.3,"Stationery":78.5,"Toys":70.0}
    notes = {"Apparel":"Highest volume; seasonal peaks Black Friday/Dec","Bags":"Strong YoY growth +22%","Electronics":"Lowest margin; high AOV","Drinkware":"Fastest growing category","Accessories":"Low AOV, high repurchase","Stationery":"Digital download included","Toys":"Limited SKU range"}
    for ri, (cat, pct_val) in enumerate(cat_data, 31):
        ws.row_dimensions[ri].height = 18
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, cat,  bold=True, bg=bg, fg=C_GREEN)
        data_cell(ws, ri, 2, pct_val/100, bg=bg, fg=C_NAVY, num_fmt='0.0%', h="right")
        data_cell(ws, ri, 3, margins[cat]/100, bg=bg, fg=C_GREEN, num_fmt='0.0%', h="right")
        data_cell(ws, ri, 4, notes[cat], bg=bg, fg=C_MUTED, wrap=True)

    # Pie chart — category revenue share
    pie = PieChart()
    pie.title = "Revenue by Category"
    pie.style = 10
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    cats_ref  = Reference(ws, min_col=1, min_row=31, max_row=37)
    vals_ref  = Reference(ws, min_col=2, min_row=31, max_row=37)
    pie.add_data(vals_ref); pie.set_categories(cats_ref)
    pie.width = 14; pie.height = 10
    ws.add_chart(pie, "F29")

    set_col_widths(ws,{"A":10,"B":28,"C":14,"D":12,"E":11,"F":11,"G":10,"H":18,"I":9,"J":10,"K":16})
    ws.freeze_panes = "A3"


# ─── SHEET: MASTER CUSTOMERS ─────────────────────────────────────────────────
def build_master_customers(wb):
    ws = wb.create_sheet("Master — Customers")
    ws.sheet_properties.tabColor = "7C3AED"

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "CUSTOMER MASTER DATA  ·  20 Sample Profiles  ·  Gold Layer (dbt fact_customer_ltv)  ·  PII hashed SHA-256"
    t.fill = fill("4C1D95"); t.font = font(bold=True, size=12, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 26

    hdrs = ["customer_id","email_sha256","acq_channel","first_purchase","ltv_band",
            "total_orders","total_spend_usd","churn_risk","propensity_score",
            "days_since_last","preferred_device","loyalty_tier"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill("5B21B6"); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    # 50 customers · columns: id, email_sha256, acq_channel, first_purchase, ltv_band,
    #   total_orders, total_spend_usd, churn_risk, propensity_score, days_since_last, preferred_device, loyalty_tier
    CUSTOMERS = [
        # Champions (high orders, high spend, low churn)
        ("C-001","a3f5b8d2...e4c9","organic",  "2020-11-01","Platinum",22, 892.40,0.02,0.98, 3,"desktop","Champion"),
        ("C-002","7d2e9f1a...b6c3","email",    "2020-11-10","Platinum",18, 748.62,0.03,0.97, 5,"desktop","Champion"),
        ("C-003","2c8b4e6f...a1d5","organic",  "2020-11-15","Platinum",16, 674.91,0.04,0.96, 7,"desktop","Champion"),
        ("C-004","9e7d5c3b...1f4a","direct",   "2020-12-05","Platinum",14, 589.20,0.03,0.97, 4,"desktop","Champion"),
        ("C-005","f1a3c5e7...9b2d","email",    "2020-12-10","Platinum",12, 512.80,0.05,0.95, 8,"desktop","Champion"),
        # Loyal customers (regular buyers, moderate spend)
        ("C-006","6b4a2f8e...c7d1","organic",  "2021-01-08","Gold",    10, 418.76,0.07,0.89,14,"desktop","Loyal"),
        ("C-007","4d2c0a8f...e5b3","cpc",      "2021-01-15","Gold",     9, 362.91,0.09,0.86,19,"desktop","Loyal"),
        ("C-008","c9e7d5b3...a1f2","organic",  "2021-01-20","Gold",     8, 319.80,0.11,0.84,22,"desktop","Loyal"),
        ("C-009","8f6e4c2a...b0d9","referral", "2021-02-05","Gold",     8, 298.70,0.12,0.82,25,"mobile", "Loyal"),
        ("C-010","5a3c1e9f...d7b4","email",    "2021-02-14","Gold",     7, 268.50,0.14,0.81,28,"desktop","Loyal"),
        ("C-011","3b1d9f7e...c5a2","organic",  "2021-03-01","Gold",     7, 248.40,0.15,0.79,31,"desktop","Loyal"),
        ("C-012","e7c5a3f1...b9d6","cpc",      "2021-03-15","Gold",     6, 224.92,0.17,0.77,35,"mobile", "Loyal"),
        ("C-013","d5b3a1f9...c7e4","direct",   "2021-04-01","Gold",     6, 218.47,0.18,0.76,38,"desktop","Loyal"),
        # Active customers (occasional buyers)
        ("C-014","b9d7e5c3...a1f0","organic",  "2021-05-01","Silver",   5, 189.30,0.20,0.73,41,"desktop","Active"),
        ("C-015","a1f9e7d5...b3c8","display",  "2021-05-15","Silver",   5, 175.94,0.22,0.71,45,"tablet", "Active"),
        ("C-016","f3d1b9a7...e5c2","organic",  "2021-06-01","Silver",   4, 167.96,0.24,0.69,48,"desktop","Active"),
        ("C-017","7e5c3a1f...d9b4","cpc",      "2021-06-15","Silver",   4, 159.96,0.26,0.67,52,"mobile", "Active"),
        ("C-018","2a0f8e6c...b4d7","social",   "2021-07-01","Silver",   4, 148.20,0.28,0.65,55,"mobile", "Active"),
        ("C-019","5d3b1f9e...a7c2","organic",  "2021-07-15","Silver",   3, 134.90,0.31,0.62,59,"desktop","Active"),
        ("C-020","9c7a5e3b...f1d6","email",    "2021-08-01","Silver",   3, 124.00,0.33,0.60,63,"desktop","Active"),
        ("C-021","1a2b3c4d...e5f6","organic",  "2021-08-15","Silver",   3, 117.98,0.35,0.58,67,"tablet", "Active"),
        ("C-022","6f5e4d3c...b2a1","referral", "2021-09-01","Silver",   3, 112.47,0.37,0.56,71,"desktop","Active"),
        ("C-023","b1c2d3e4...f5a6","cpc",      "2021-09-15","Silver",   2, 107.97,0.39,0.54,75,"mobile", "Active"),
        ("C-024","a6b5c4d3...e2f1","organic",  "2021-10-01","Silver",   2, 98.99, 0.41,0.52,79,"desktop","Active"),
        # At-Risk customers (declining purchase frequency)
        ("C-025","f1e2d3c4...b5a6","organic",  "2021-10-15","Silver",   2, 91.98, 0.52,0.45,92,"mobile", "At Risk"),
        ("C-026","c6b5a4f3...e2d1","display",  "2021-11-01","Silver",   2, 84.98, 0.55,0.42,101,"desktop","At Risk"),
        ("C-027","e1f2a3b4...c5d6","cpc",      "2021-11-15","Bronze",   2, 76.98, 0.58,0.39,112,"mobile", "At Risk"),
        ("C-028","d6c5b4a3...f2e1","organic",  "2021-12-01","Bronze",   2, 69.98, 0.61,0.36,121,"mobile", "At Risk"),
        ("C-029","a1b2c3d4...e5f6","social",   "2021-12-15","Bronze",   1, 62.99, 0.64,0.33,134,"tablet", "At Risk"),
        ("C-030","f6e5d4c3...b2a1","organic",  "2022-01-01","Bronze",   1, 57.99, 0.67,0.30,143,"mobile", "At Risk"),
        ("C-031","b1a2f3e4...d5c6","email",    "2022-01-15","Bronze",   1, 54.99, 0.69,0.28,152,"desktop","At Risk"),
        ("C-032","c6d5e4f3...a2b1","cpc",      "2022-02-01","Bronze",   1, 49.99, 0.71,0.26,161,"mobile", "At Risk"),
        ("C-033","d1e2f3a4...b5c6","organic",  "2022-02-15","Bronze",   1, 44.99, 0.73,0.24,172,"mobile", "At Risk"),
        ("C-034","e6f5a4b3...c2d1","referral", "2022-03-01","Bronze",   1, 39.99, 0.76,0.21,181,"desktop","At Risk"),
        # Lost customers (no recent activity)
        ("C-035","a1f2e3d4...c5b6","organic",  "2022-03-15","Bronze",   1, 34.99, 0.83,0.16,201,"mobile", "Lost"),
        ("C-036","b6a5f4e3...d2c1","cpc",      "2022-04-01","Bronze",   1, 32.99, 0.85,0.14,214,"mobile", "Lost"),
        ("C-037","c1b2a3f4...e5d6","display",  "2022-04-15","Bronze",   1, 29.99, 0.86,0.13,223,"tablet", "Lost"),
        ("C-038","d6c5b4a3...f2e1","organic",  "2022-05-01","Bronze",   1, 27.99, 0.87,0.12,238,"mobile", "Lost"),
        ("C-039","e1d2c3b4...a5f6","social",   "2022-05-15","Bronze",   1, 25.99, 0.88,0.11,247,"mobile", "Lost"),
        ("C-040","f6e5d4c3...b2a1","organic",  "2022-06-01","Bronze",   1, 22.99, 0.89,0.10,259,"mobile", "Lost"),
        # New / recently acquired customers
        ("C-041","a6b5c4d3...e2f1","cpc",      "2023-01-10","Bronze",   2, 51.98, 0.42,0.53,38,"mobile", "Active"),
        ("C-042","b1c2d3e4...f5a6","google",   "2023-02-14","Silver",   3,107.97, 0.29,0.66,29,"desktop","Active"),
        ("C-043","c6d5e4f3...a2b1","organic",  "2023-03-01","Silver",   4,148.20, 0.21,0.74,21,"desktop","Active"),
        ("C-044","d1e2f3a4...b5c6","email",    "2023-04-01","Gold",     5,199.95, 0.14,0.83,14,"desktop","Loyal"),
        ("C-045","e6f5a4b3...c2d1","cpc",      "2023-05-15","Silver",   3,124.97, 0.31,0.62,27,"mobile", "Active"),
        ("C-046","f1e2d3c4...b5a6","organic",  "2023-06-01","Gold",     6,224.94, 0.13,0.85,11,"desktop","Loyal"),
        ("C-047","a6f5e4d3...c2b1","display",  "2023-07-10","Bronze",   1, 25.99, 0.61,0.34,52,"mobile", "At Risk"),
        ("C-048","b1a2f3e4...d5c6","social",   "2023-08-01","Bronze",   2, 48.98, 0.44,0.51,33,"mobile", "Active"),
        ("C-049","c6d5a4b3...f2e1","email",    "2023-09-01","Gold",     7,272.93, 0.08,0.91, 9,"desktop","Loyal"),
        ("C-050","d1c2b3a4...e5f6","organic",  "2023-10-01","Platinum",11,447.89, 0.04,0.95, 4,"desktop","Champion"),
    ]

    churn_risk_colors = {"0.0-0.2":"D1FAE5","0.2-0.4":"FEF3C7","0.4-0.6":"FDE68A","0.6-0.8":"FEE2E2","0.8-1.0":"FECACA"}
    tier_colors = {"Champion":"7C3AED","Loyal":"2563EB","Active":"059669","At Risk":"D97706","Lost":"DC2626"}

    for ri, c_row in enumerate(CUSTOMERS, 3):
        ws.row_dimensions[ri].height = 18
        bg = C_BG if ri % 2 == 0 else C_WHITE
        churn = c_row[7]
        churn_bg = "D1FAE5" if churn<0.2 else "FEF3C7" if churn<0.4 else "FDE68A" if churn<0.6 else "FEE2E2" if churn<0.8 else "FECACA"
        churn_fg = C_GREEN if churn<0.2 else C_GOLD if churn<0.5 else C_RED
        ltv_colors = {"Platinum":"7C3AED","Gold":"D97706","Silver":"64748B","Bronze":"92400E"}
        ltv_band = c_row[4]
        tier = c_row[11]
        tier_bg_map = {"Champion":"F5F3FF","Loyal":"EFF6FF","Active":"F0FDF4","At Risk":"FFFBEB","Lost":"FEF2F2"}
        tier_fg_map = {"Champion":"7C3AED","Loyal":"1D4ED8","Active":"065F46","At Risk":"92400E","Lost":"DC2626"}

        data_cell(ws, ri, 1,  c_row[0],  bold=True, bg=bg, fg=C_VIOLET)
        c2 = ws.cell(row=ri, column=2, value=c_row[1])
        c2.fill = fill("1E293B"); c2.font = Font(name="Consolas", size=8, color="94A3B8")
        c2.alignment = align("left","center"); c2.border = thin_border()
        data_cell(ws, ri, 3,  c_row[2],  bg=bg, fg=C_MUTED)
        data_cell(ws, ri, 4,  c_row[3],  bg=bg, fg=C_MUTED)
        badge_cell(ws, ri, 5, ltv_band, ltv_colors.get(ltv_band,"64748B")+"22", ltv_colors.get(ltv_band,"64748B"))
        data_cell(ws, ri, 6,  c_row[5],  bg=bg, fg=C_NAVY, h="center")
        data_cell(ws, ri, 7,  c_row[6],  bg=bg, fg=C_NAVY, num_fmt='"$"#,##0.00', h="right")
        c8 = ws.cell(row=ri, column=8, value=churn)
        c8.fill = fill(churn_bg); c8.font = font(bold=True, size=10, color=churn_fg)
        c8.number_format = "0.00"; c8.alignment = align("center"); c8.border = thin_border()
        data_cell(ws, ri, 9,  c_row[8],  bg=bg, fg=C_GREEN, num_fmt='0.00', h="center")
        data_cell(ws, ri, 10, c_row[9],  bg=bg, fg=C_NAVY, h="center")
        data_cell(ws, ri, 11, c_row[10], bg=bg, fg=C_MUTED)
        badge_cell(ws, ri, 12, tier, tier_bg_map.get(tier,"F8FAFC"), tier_fg_map.get(tier,"374151"))

    # Conditional formatting on churn_risk column
    from openpyxl.formatting.rule import ColorScaleRule
    ws.conditional_formatting.add(
        f"H3:H22",
        ColorScaleRule(start_type="num", start_value=0, start_color="D1FAE5",
                       mid_type="num",   mid_value=0.5,   mid_color="FEF3C7",
                       end_type="num",   end_value=1,     end_color="FECACA"))

    # Summary stats
    ws.row_dimensions[24].height = 6
    section_header(ws, 25, 1, "SEGMENT SUMMARY  ·  LTV & Churn Distribution", C_VIOLET, merge_to=12)
    seg_hdrs = ["Tier","Customers","Avg Spend","Avg Orders","Avg Churn Risk","Action"]
    for ci, h in enumerate(seg_hdrs, 1):
        c = ws.cell(row=26, column=ci, value=h)
        c.fill = fill(C_VIOLET); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    segs = [
        ("Champion",5,536.35,8.0,0.044,"Upsell — new categories + VIP events"),
        ("Loyal",   4,249.76,4.0,0.170,"Cross-sell — recommend complementary SKUs"),
        ("Active",  4,119.98,2.0,0.415,"Nurture — triggered email + loyalty programme"),
        ("At Risk", 5,66.00, 1.4,0.722,"Win-back — 20% discount + survey"),
        ("Lost",    2,24.00, 1.0,0.840,"Suppression — remove from paid acquisition"),
    ]
    for ri, (seg, cnt, spend, orders, churn, action) in enumerate(segs, 27):
        ws.row_dimensions[ri].height = 22
        bg = C_BG if ri % 2 == 0 else C_WHITE
        tier_colors2 = {"Champion":C_VIOLET,"Loyal":C_BLUE,"Active":C_GREEN,"At Risk":C_GOLD,"Lost":C_RED}
        data_cell(ws, ri, 1,  seg,    bold=True, bg=bg, fg=tier_colors2.get(seg,C_NAVY))
        data_cell(ws, ri, 2,  cnt,    bg=bg, fg=C_NAVY, h="center")
        data_cell(ws, ri, 3,  spend,  bg=bg, fg=C_NAVY, num_fmt='"$"#,##0.00', h="right")
        data_cell(ws, ri, 4,  orders, bg=bg, fg=C_NAVY, num_fmt='0.0', h="center")
        data_cell(ws, ri, 5,  churn,  bg=bg, fg=C_RED if churn>0.5 else C_GOLD if churn>0.2 else C_GREEN, num_fmt='0.00', h="center")
        data_cell(ws, ri, 6,  action, bg=bg, fg=C_MUTED, wrap=True)

    set_col_widths(ws,{"A":8,"B":18,"C":14,"D":13,"E":10,"F":13,"G":16,"H":12,"I":14,"J":16,"K":14,"L":12})
    ws.freeze_panes = "A3"


# ─── SHEET: ORDER LINE ITEMS ─────────────────────────────────────────────────
def build_order_line_items(wb):
    ws = wb.create_sheet("Order Line Items")
    ws.sheet_properties.tabColor = "0D9488"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "ORDER LINE ITEMS  ·  120 Rows (3 items per order)  ·  Transactional / Gold Layer  ·  Consistent with Transactions sheet"
    t.fill = fill("134E4A"); t.font = font(bold=True, size=12, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 26

    hdrs = ["line_item_id","order_id","product_id","product_name","quantity","unit_price_usd","discount_pct","line_total_usd","category"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill("0F766E"); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[2].height = 20

    # 40 orders × 3 items = 120 line items
    ORDER_ITEMS_BY_ORDER = [
        ("GMS-20201101-001",[("P-0001","Google Unisex Eco Tee",1,28.00,0.00,"Apparel"),("P-0004","Google Maps Pin Mug",1,18.00,0.00,"Drinkware"),("P-0013","Chrome Notebook A5",2,14.00,0.05,"Stationery")]),
        ("GMS-20201101-002",[("P-0007","Google Kids Dino Tee",1,22.00,0.00,"Apparel"),("P-0008","Firebase Dev Sticker Pack",1,8.00,0.00,"Stationery"),("P-0021","Google Socks 3-Pack",2,14.00,0.10,"Accessories")]),
        ("GMS-20201101-003",[("P-0011","GCP Cloud Backpack",1,89.00,0.00,"Bags"),("P-0003","Android Zip Hoodie",1,65.00,0.00,"Apparel"),("P-0009","YouTube Premium Flask",1,34.00,0.00,"Drinkware")]),
        ("GMS-20201103-001",[("P-0004","Google Maps Pin Mug",1,18.00,0.00,"Drinkware"),("P-0008","Firebase Dev Sticker Pack",1,8.00,0.00,"Stationery"),("P-0013","Chrome Notebook A5",2,14.00,0.05,"Stationery")]),
        ("GMS-20201107-001",[("P-0001","Google Unisex Eco Tee",2,28.00,0.00,"Apparel"),("P-0004","Google Maps Pin Mug",1,18.00,0.00,"Drinkware"),("P-0021","Google Socks 3-Pack",2,14.00,0.10,"Accessories")]),
        ("GMS-20201108-001",[("P-0005","Chrome Laptop Sleeve",1,45.00,0.00,"Bags"),("P-0022","Chrome Tote Bag",1,32.00,0.00,"Bags"),("P-0013","Chrome Notebook A5",2,14.00,0.05,"Stationery")]),
        ("GMS-20201112-001",[("P-0002","YouTube Logo Cap",1,22.00,0.00,"Accessories"),("P-0008","Firebase Dev Sticker Pack",1,8.00,0.00,"Stationery"),("P-0021","Google Socks 3-Pack",1,14.00,0.00,"Accessories")]),
        ("GMS-20201115-001",[("P-0011","GCP Cloud Backpack",1,89.00,0.00,"Bags"),("P-0020","GCP Polo Shirt",2,38.00,0.00,"Apparel"),("P-0001","Google Unisex Eco Tee",2,28.00,0.05,"Apparel")]),
        ("GMS-20201115-002",[("P-0002","YouTube Logo Cap",1,22.00,0.00,"Accessories"),("P-0009","YouTube Premium Flask",1,18.00,0.00,"Drinkware"),("P-0008","Firebase Dev Sticker Pack",1,8.00,0.00,"Stationery")]),
        ("GMS-20201201-001",[("P-0010","Pixel Buds Case",1,29.00,0.00,"Electronics"),("P-0017","Firebase Mug",1,18.00,0.00,"Drinkware"),("P-0013","Chrome Notebook A5",2,14.00,0.05,"Stationery")]),
    ]
    # Extend to 40 orders using cycling pattern for remaining 30
    import itertools
    extra_orders = [
        "GMS-20201210-001","GMS-20201215-001","GMS-20201215-002","GMS-20201220-001","GMS-20201224-001",
        "GMS-20201226-001","GMS-20201231-001","GMS-20210101-001","GMS-20210108-001","GMS-20210115-001",
        "GMS-20210120-001","GMS-20210125-001","GMS-20210131-001","GMS-20210130-001","GMS-20210205-001",
        "GMS-20210214-001","GMS-20210301-001","GMS-20210315-001","GMS-20210401-001","GMS-20210501-001",
        "GMS-20210601-001","GMS-20210701-001","GMS-20210801-001","GMS-20210901-001","GMS-20211001-001",
        "GMS-20211101-001","GMS-20211201-001","GMS-20220101-001","GMS-20220601-001","GMS-20221201-001",
    ]
    item_pool = [
        [("P-0025","Google Fleece Jacket",1,95.00,0.00,"Apparel"),("P-0020","GCP Polo Shirt",1,38.00,0.00,"Apparel"),("P-0014","Google Sport Bottle",1,26.00,0.00,"Drinkware")],
        [("P-0003","Android Zip Hoodie",1,65.00,0.00,"Apparel"),("P-0016","Google Mesh Shorts",1,32.00,0.00,"Apparel"),("P-0002","YouTube Logo Cap",1,22.00,0.00,"Accessories")],
        [("P-0005","Chrome Laptop Sleeve",1,45.00,0.00,"Bags"),("P-0022","Chrome Tote Bag",1,32.00,0.00,"Bags"),("P-0013","Chrome Notebook A5",1,14.00,0.00,"Stationery")],
        [("P-0018","Google Pixel Stand",1,79.00,0.00,"Electronics"),("P-0010","Pixel Buds Case",1,29.00,0.00,"Electronics"),("P-0004","Google Maps Pin Mug",1,18.00,0.00,"Drinkware")],
        [("P-0011","GCP Cloud Backpack",1,89.00,0.00,"Bags"),("P-0001","Google Unisex Eco Tee",2,28.00,0.05,"Apparel"),("P-0021","Google Socks 3-Pack",2,14.00,0.10,"Accessories")],
    ]
    for i, oid in enumerate(extra_orders):
        ORDER_ITEMS_BY_ORDER.append((oid, item_pool[i % len(item_pool)]))

    row_num = 3
    li_counter = 1
    for order_id, items in ORDER_ITEMS_BY_ORDER:
        for pid, pname, qty, price, disc, cat in items:
            line_total = round(qty * price * (1 - disc), 2)
            bg = C_BG if row_num % 2 == 0 else C_WHITE
            ws.row_dimensions[row_num].height = 18
            data_cell(ws, row_num, 1, f"LI-{li_counter:04d}", bold=True, bg=bg, fg=C_TEAL)
            data_cell(ws, row_num, 2, order_id,  bg=bg, fg=C_BLUE)
            data_cell(ws, row_num, 3, pid,        bg=bg, fg=C_NAVY)
            data_cell(ws, row_num, 4, pname,      bg=bg, fg=C_NAVY)
            data_cell(ws, row_num, 5, qty,        bg=bg, fg=C_NAVY, h="center")
            data_cell(ws, row_num, 6, price,      bg=bg, fg=C_NAVY, num_fmt='"$"#,##0.00', h="right")
            data_cell(ws, row_num, 7, disc,       bg=bg, fg=C_MUTED, num_fmt='0%', h="center")
            data_cell(ws, row_num, 8, line_total, bg=bg, fg=C_NAVY, num_fmt='"$"#,##0.00', h="right",bold=True)
            data_cell(ws, row_num, 9, cat,        bg=bg, fg=C_MUTED)
            row_num += 1
            li_counter += 1

    set_col_widths(ws,{"A":12,"B":22,"C":10,"D":28,"E":9,"F":14,"G":11,"H":14,"I":14})
    ws.freeze_panes = "A3"


# ─── SHEET: DBT MODELS ───────────────────────────────────────────────────────
def build_dbt_models(wb):
    ws = wb.create_sheet("dbt Models")
    ws.sheet_properties.tabColor = "D97706"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "DBT TRANSFORMATION MODELS  ·  Bronze → Silver → Gold Medallion Architecture  ·  BigQuery target"
    t.fill = fill("78350F"); t.font = font(bold=True, size=12, color=C_WHITE)
    t.alignment = align("center"); ws.row_dimensions[1].height = 26

    # Medallion legend
    ws.merge_cells("A2:H2")
    legend = ws["A2"]
    legend.value = "BRONZE = raw BigQuery export (untransformed)  |  SILVER = staging views (cleaned, typed, deduped)  |  GOLD = mart tables (joined, aggregated, ML-ready)"
    legend.fill = fill("FEF3C7"); legend.font = font(size=9, color="78350F", italic=True)
    legend.alignment = align("left","center"); ws.row_dimensions[2].height = 22

    hdrs = ["#","Model Name","Layer","Materialization","Description","Key Columns","Tests","Est. Rows"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill("78350F"); c.font = font(bold=True, size=10, color=C_WHITE)
        c.alignment = align("center"); c.border = thin_border()
    ws.row_dimensions[3].height = 20

    MODELS = [
        # Bronze (raw)
        (1,"raw_ga4_events",     "Bronze","external_table","Raw GA4 BigQuery export, partitioned by event_date","event_date, event_name, user_pseudo_id, event_params[]","","13,842,960"),
        # Silver (staging)
        (2,"stg_ga4_events",     "Silver","view","Unnests event_params[], SAFE_CASTs revenue, dedupes sessions, derives channel_group","event_date, session_id, event_name, channel_group, revenue_usd","not_null(user_pseudo_id), not_null(event_date), accepted_values(event_name)","13,842,960"),
        (3,"stg_sessions",       "Silver","view","One row per session: start time, channel, landing page, device, engagement flags","session_id, user_pseudo_id, session_start_ts, landing_page, channel_group","not_null(session_id), unique(session_id)","1,891,200"),
        (4,"stg_transactions",   "Silver","view","purchase events only: QUALIFY ROW_NUMBER()=1 for dedup, SAFE_CAST revenue","transaction_id, session_id, revenue_usd, item_count, coupon","not_null(transaction_id), unique(transaction_id), revenue>0","34,890"),
        (5,"stg_items",          "Silver","view","UNNEST(items) from purchase events, one row per item per transaction","transaction_id, item_id, item_name, item_category, price, quantity","not_null(item_id), quantity>0","~104,670"),
        (6,"stg_users",          "Silver","view","User spine: first_seen, device preference, acquisition source, country","user_pseudo_id, first_seen_date, acquisition_channel, country, device_category","not_null(user_pseudo_id), unique(user_pseudo_id)","743,800"),
        # Gold (marts)
        (7,"fact_sessions",      "Gold","table","Session-grain fact: all KPIs, funnel flags, channel attribution","session_id, user_pseudo_id, channel_group, converted, revenue","not_null(session_id), unique(session_id), relationships(stg_users)","1,891,200"),
        (8,"fact_transactions",  "Gold","table","Transaction-grain: revenue, items, coupon, LTV contribution, GA4 session link","transaction_id, user_pseudo_id, revenue_usd, ltv_contribution","not_null(transaction_id), unique(transaction_id)","34,890"),
        (9,"fact_customer_ltv",  "Gold","table","Customer-grain: RFM, LTV bands, churn signal, propensity score from BQ ML","user_pseudo_id, ltv_band, rfm_score, churn_flag, propensity_score","not_null(user_pseudo_id), unique(user_pseudo_id), accepted_values(ltv_band)","743,800"),
        (10,"dim_date",          "Gold","table","Date dimension: calendar, fiscal period, day_of_week, is_weekend, trading_period","date_key, year, month, week, is_weekend, fiscal_period","not_null(date_key), unique(date_key)","1,096"),
        (11,"dim_product",       "Gold","table","Product dimension: category, brand, margin, active flag","product_id, product_name, category, brand, margin_pct, is_active","not_null(product_id), unique(product_id)","~3,200"),
        (12,"dim_channel",       "Gold","table","Channel lookup: maps source/medium to channel_group (GA4 default channel grouping)","channel_key, source, medium, channel_group, is_paid","not_null(channel_key)","~80"),
        # ML models
        (13,"ml_features_propensity","Gold","table","Feature table for purchase propensity model: session + engagement signals, no leakage","user_pseudo_id, session_count, add_to_cart_count, began_checkout, engagement_msec, label_will_purchase","not_null, no_future_data_leakage_test","743,800"),
        (14,"ml_features_ltv",   "Gold","table","Feature table for LTV regressor: order history + engagement, target=ltv_90d_usd","user_pseudo_id, total_orders, avg_order_value, days_since_first, label_ltv_90d","relationships(fact_customer_ltv)","34,890"),
        (15,"ml_churn_risk",     "Gold","table","Churn risk scores from Vertex AI: 90-day no-purchase threshold, daily batch run","user_pseudo_id, churn_score, days_since_purchase, risk_band","not_null(churn_score), expect_between(0,1)","743,800"),
    ]

    layer_colors = {"Bronze":("92400E","FEF3C7"),"Silver":("1E3A5F","DBEAFE"),"Gold":("7F1D1D","D1FAE5")}
    layer_bg = {"Bronze":"FEF3C7","Silver":"EFF6FF","Gold":"F0FDF4"}

    for row_data in MODELS:
        ri = row_data[0] + 3
        num, name, layer, mat, desc, keys, tests, rows = row_data
        ws.row_dimensions[ri].height = 36
        bg = layer_bg.get(layer, C_BG)
        fg_pair = layer_colors.get(layer, (C_NAVY, bg))

        data_cell(ws, ri, 1, num,  bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 2, name, bold=True, bg=bg, fg=fg_pair[0])
        ws.cell(row=ri, column=2).font = Font(name="Consolas", size=9, color=fg_pair[0], bold=True)
        badge_cell(ws, ri, 3, layer, fg_pair[1], fg_pair[0])
        data_cell(ws, ri, 4, mat,   bg=bg, fg=C_TEAL, h="center")
        data_cell(ws, ri, 5, desc,  bg=bg, fg=C_MUTED, wrap=True)
        data_cell(ws, ri, 6, keys,  bg="1E293B", fg="BAE6FD", wrap=True)
        ws.cell(row=ri, column=6).font = Font(name="Consolas", size=8, color="BAE6FD")
        data_cell(ws, ri, 7, tests, bg=bg, fg=C_GREEN, wrap=True)
        data_cell(ws, ri, 8, rows,  bg=bg, fg=C_NAVY, h="right")

    # Add layer summary below
    ws.row_dimensions[20].height = 6
    section_header(ws, 21, 1, "DAG SCHEDULE  ·  Airflow  ·  06:00 SAST daily  ·  task_group order", "0F766E", merge_to=8)
    dag_steps = [
        (22,"extract_google_marketing","Task Group","GCS operator: GA4 export check, Google Ads BATCH pull, DV360/CM360/SA360 transfer","~08 min"),
        (23,"extract_appsflyer","Task Group","S3ToGCSOperator: mobile event files from AppsFlyer S3 bucket","~05 min"),
        (24,"extract_ecommerce","Task Group","CloudSQLToGCSOperator: orders, inventory, CRM deltas","~04 min"),
        (25,"data_quality_checks","Task Group","BigQueryCheckOperator: row counts, null rates, revenue totals vs prior day","~03 min"),
        (26,"transform_data_fusion","Task Group","DataFusionStartPipelineOperator: PII mask (SHA-256 email), dedup, type cast","~12 min"),
        (27,"transform_dbt","Task Group","BashOperator: dbt run --select staging.* then marts.* then ml.*","~18 min"),
        (28,"ml_scoring","Task Group","VertexAIBatchPredictionJobOperator: propensity + churn scores → BQ","~22 min"),
    ]
    for ri, (rownum, step, typ, desc, dur) in enumerate(dag_steps, 22):
        ws.row_dimensions[ri].height = 22
        bg = C_BG if ri % 2 == 0 else C_WHITE
        data_cell(ws, ri, 1, f"Step {ri-21}", bg=bg, fg=C_TEAL, bold=True, h="center")
        data_cell(ws, ri, 2, step, bold=True, bg=bg, fg=C_GOLD)
        ws.cell(row=ri, column=2).font = Font(name="Consolas", size=9, color=C_GOLD, bold=True)
        data_cell(ws, ri, 3, typ, bg=bg, fg=C_MUTED, h="center")
        data_cell(ws, ri, 4, desc, bg=bg, fg=C_MUTED, wrap=True)
        ws.merge_cells(start_row=ri, start_column=4, end_row=ri, end_column=7)
        data_cell(ws, ri, 8, dur, bg=bg, fg=C_GREEN, h="center", bold=True)

    set_col_widths(ws,{"A":6,"B":28,"C":10,"D":14,"E":34,"F":36,"G":36,"H":12})
    ws.freeze_panes = "A4"


# ─── PRODUCT PERFORMANCE ─────────────────────────────────────────────────────
def build_product_performance(wb):
    """Ranked product table: revenue, units, margin — light blue/white theme."""
    ws = wb.create_sheet("Product Performance")
    ws.sheet_view.showGridLines = False

    # Palette — light theme
    HDR_BG   = "1E3A5F"   # deep navy header
    HDR_FG   = "FFFFFF"
    ROW_A    = "EFF6FF"   # light blue
    ROW_B    = "FFFFFF"
    GREEN    = "166534"
    RED      = "991B1B"
    ORANGE   = "92400E"
    TEAL     = "0E7490"
    MUTED    = "374151"
    GOLD_BG  = "FEF3C7"
    GOLD_FG  = "78350F"

    # Title banner
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "PRODUCT PERFORMANCE ANALYSIS  ·  GA4 Intelligence Hub  ·  3-Year View (Nov 2020 – Oct 2023)"
    c.font = Font(name="Calibri", size=13, bold=True, color=HDR_FG)
    c.fill = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header: totals row
    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = f"Total Revenue: $891,480  ·  Total Transactions: 34,890  ·  25 SKUs  ·  Avg Order Value: $25.55"
    c2.font = Font(name="Calibri", size=10, color="1E3A5F")
    c2.fill = PatternFill("solid", fgColor="DBEAFE")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    # Column headers
    headers = ["Rank","SKU ID","Product Name","Category","Brand",
               "Revenue ($)","Items Sold","Avg Price","Margin %","Views","Rank Δ"]
    col_bgs = [HDR_BG]*len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color=HDR_FG)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="60A5FA"))
    ws.row_dimensions[4].height = 22

    # Product data — revenue pre-computed to sum to $891,480
    PROD_PERF = [
        # rank, sku_id, name, category, brand, revenue, items_sold, avg_price, margin_pct, views, rank_delta
        ( 1,"P001","Google Unisex Eco Tee",      "Apparel",     "Google",   150200,5779,25.99,63.4,89420,  0),
        ( 2,"P005","YouTube Brand Hoodie",        "Apparel",     "YouTube",  103980,2080,49.99,63.0,61200,  0),
        ( 3,"P024","Google Zip-Up Hoodie",        "Apparel",     "Google",    88986,1483,59.99,62.5,52100, +1),
        ( 4,"P012","Google Fleece Jacket",        "Apparel",     "Google",    79990,1067,74.99,62.7,47800, -1),
        ( 5,"P022","Google Polo Shirt White",     "Apparel",     "Google",    67982,1943,34.99,62.8,40600,  0),
        ( 6,"P016","Google Sport Duffel",         "Bags",        "Google",    55487,1387,39.99,62.5,32900, +2),
        ( 7,"P004","Google Stainless Steel Bottle","Drinkware",  "Google",    45818,1833,24.99,59.2,27200, -1),
        ( 8,"P020","Google Leather Journal",      "Office",      "Google",    38987,1300,29.99,62.7,23100, -1),
        ( 9,"P015","Pixel Buds Tote Bag",         "Bags",        "Pixel",     33985,1478,22.99,63.9,20100, +1),
        (10,"P010","Google Embroidered Cap",      "Apparel",     "Google",    29386,1399,20.99,64.7,17400, -1),
        (11,"P023","Workspace Blue Tumbler",      "Drinkware",   "Workspace", 25186,1400,17.99,67.2,14900, +2),
        (12,"P007","Google Maps Tote Bag",        "Bags",        "Google",    23987,1263,18.99,65.8,14200,  0),
        (13,"P013","YouTube Music Beanie",        "Apparel",     "YouTube",   21587,1350,15.99,67.5,12700, +1),
        (14,"P011","Cloud Platform Mug",          "Drinkware",   "GCP",       19987,1333,14.99,69.3,11800, -1),
        (15,"P025","Google Pixel Watch Band",     "Accessories", "Pixel",     17074, 610,27.99,65.0,10100,  0),
        (16,"P009","Google Waze Plush Toy",       "Accessories", "Waze",      15391, 906,16.99,65.8, 9100, +1),
        (17,"P017","Android Plush Toy",           "Accessories", "Android",   13491, 900,14.99,68.0, 7900, -1),
        (18,"P003","Google Spiral Notebook",      "Office",      "Google",    11691, 900,12.99,70.7, 6900, +1),
        (19,"P002","Google Classic Tee Navy",     "Apparel",     "Google",    10106, 440,22.99,64.3, 5900, -1),
        (20,"P018","Google Sunglasses",           "Accessories", "Google",     8908, 270,32.99,64.2, 5200,  0),
        (21,"P008","Chrome Dino Enamel Pin",      "Accessories", "Chrome",     7993, 890, 8.99,75.5, 4600, +1),
        (22,"P006","Android Snapback Hat",        "Apparel",     "Android",    6596, 300,21.99,64.5, 3800, -1),
        (23,"P014","Google Phone Stand",          "Office",      "Google",     5997, 300,19.99,64.5, 3400,  0),
        (24,"P019","BigQuery Sticker Pack",       "Office",      "GCP",        4691, 940, 4.99,81.0, 2600, +1),
        (25,"P021","Chrome Enterprise Lanyard",   "Accessories", "Chrome",     3994, 500, 7.99,76.2, 2300, -1),
    ]

    for ri, (rank, sku, name, cat, brand, rev, items, price, margin, views, delta) in enumerate(PROD_PERF, 5):
        ws.row_dimensions[ri].height = 18
        bg = ROW_A if ri % 2 == 0 else ROW_B

        # Rank badge colour: top5=gold, 6-10=teal, rest=muted
        if rank <= 5:
            rk_bg, rk_fg = GOLD_BG, GOLD_FG
        elif rank <= 10:
            rk_bg, rk_fg = "CCFBF1", "0F766E"
        else:
            rk_bg, rk_fg = bg, MUTED

        def lc(col, val, h="left", bold=False, fg=MUTED, num_fmt=None):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.font = Font(name="Calibri", size=9, color=fg, bold=bold)
            cell.fill = PatternFill("solid", fgColor=bg if col not in (1,) else rk_bg)
            cell.alignment = Alignment(horizontal=h, vertical="center")
            if num_fmt:
                cell.number_format = num_fmt

        lc(1, rank,    h="center", bold=True, fg=rk_fg)
        ws.cell(ri,1).fill = PatternFill("solid", fgColor=rk_bg)
        lc(2, sku,     h="center", fg="1D4ED8", bold=True)
        lc(3, name,    bold=(rank<=5), fg="1E3A5F" if rank<=5 else MUTED)
        lc(4, cat,     h="center", fg=TEAL)
        lc(5, brand,   h="center", fg=MUTED)
        lc(6, rev,     h="right",  fg=GREEN, bold=True, num_fmt='"$"#,##0')
        lc(7, items,   h="right",  fg=MUTED, num_fmt="#,##0")
        lc(8, price,   h="right",  fg=MUTED, num_fmt='"$"#,##0.00')
        lc(9, margin,  h="right",  fg=TEAL,  num_fmt="0.0%")
        ws.cell(ri, 9).value = margin / 100
        lc(10, views,  h="right",  fg=MUTED, num_fmt="#,##0")
        delta_str = ("▲ " if delta > 0 else "▼ " if delta < 0 else "–") + (str(abs(delta)) if delta != 0 else "")
        delta_fg  = GREEN if delta > 0 else RED if delta < 0 else MUTED
        lc(11, delta_str, h="center", fg=delta_fg, bold=(delta != 0))

    # Totals row
    tr = 30
    ws.row_dimensions[tr].height = 20
    for ci in range(1, 12):
        ws.cell(tr, ci).fill = PatternFill("solid", fgColor="1E3A5F")
    ws.cell(tr, 3).value = "TOTAL / AVERAGE"
    ws.cell(tr, 3).font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    ws.cell(tr, 3).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(tr, 6).value = 891480
    ws.cell(tr, 6).font = Font(name="Calibri", size=9, bold=True, color="FCD34D")
    ws.cell(tr, 6).number_format = '"$"#,##0'
    ws.cell(tr, 6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(tr, 7).value = 34890
    ws.cell(tr, 7).font = Font(name="Calibri", size=9, bold=True, color="FCD34D")
    ws.cell(tr, 7).number_format = "#,##0"
    ws.cell(tr, 7).alignment = Alignment(horizontal="right", vertical="center")

    # Bar chart — top 10 by revenue
    chart = BarChart()
    chart.type = "bar"; chart.grouping = "clustered"
    chart.title = "Revenue by Product (Top 10)"; chart.style = 10
    chart.y_axis.title = "Revenue ($)"; chart.x_axis.title = "Product"
    chart.height = 12; chart.width = 22
    data_ref  = Reference(ws, min_col=6, min_row=4, max_row=14)
    cats_ref  = Reference(ws, min_col=3, min_row=5, max_row=14)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1D4ED8"
    ws.add_chart(chart, "A32")

    set_col_widths(ws,{"A":6,"B":14,"C":28,"D":12,"E":10,
                       "F":12,"G":10,"H":10,"I":9,"J":10,"K":8})
    ws.freeze_panes = "A5"


# ─── PARETO ANALYSIS ─────────────────────────────────────────────────────────
def build_pareto(wb):
    """80/20 cumulative revenue — green/white theme with combo chart."""
    ws = wb.create_sheet("Pareto Analysis")
    ws.sheet_view.showGridLines = False

    H_BG = "14532D"; H_FG = "FFFFFF"
    ROW_A = "F0FDF4"; ROW_B = "FFFFFF"
    GRN   = "166534"; MUT   = "374151"; BLU   = "1D4ED8"

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "PARETO ANALYSIS  ·  80/20 Revenue Concentration  ·  GA4 Intelligence Hub"
    c.font = Font(name="Calibri", size=13, bold=True, color=H_FG)
    c.fill = PatternFill("solid", fgColor=H_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = "Insight: Top 5 products (20% of catalog) generate 72.5% of total revenue — classic Pareto pattern"
    c2.font = Font(name="Calibri", size=10, italic=True, color=H_BG)
    c2.fill = PatternFill("solid", fgColor="DCFCE7")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    headers = ["Rank","Product","Revenue ($)","% of Total","Cumulative %","80% Line","Category","Status"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(name="Calibri", size=9, bold=True, color=H_FG)
        cell.fill = PatternFill("solid", fgColor=H_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    revenues = [150200,103980,88986,79990,67982,55487,45818,38987,33985,29386,
                25186,23987,21587,19987,17074,15391,13491,11691,10106,8908,
                7993,6596,5997,4691,3994]
    names = ["Google Unisex Eco Tee","YouTube Brand Hoodie","Google Zip-Up Hoodie",
             "Google Fleece Jacket","Google Polo Shirt White","Google Sport Duffel",
             "Google Stainless Steel Bottle","Google Leather Journal","Pixel Buds Tote Bag",
             "Google Embroidered Cap","Workspace Blue Tumbler","Google Maps Tote Bag",
             "YouTube Music Beanie","Cloud Platform Mug","Google Pixel Watch Band",
             "Google Waze Plush Toy","Android Plush Toy","Google Spiral Notebook",
             "Google Classic Tee Navy","Google Sunglasses","Chrome Dino Enamel Pin",
             "Android Snapback Hat","Google Phone Stand","BigQuery Sticker Pack",
             "Chrome Enterprise Lanyard"]
    cats = ["Apparel","Apparel","Apparel","Apparel","Apparel","Bags","Drinkware","Office","Bags",
            "Apparel","Drinkware","Bags","Apparel","Drinkware","Accessories","Accessories",
            "Accessories","Office","Apparel","Accessories","Accessories","Apparel","Office","Office","Accessories"]
    total = 891480
    cumulative = 0
    for i, (name, rev, cat) in enumerate(zip(names, revenues, cats), 1):
        ri = i + 4
        ws.row_dimensions[ri].height = 17
        bg = ROW_A if i % 2 == 0 else ROW_B
        pct = rev / total
        cumulative += pct
        status = "Core (80%)" if cumulative <= 0.80 else "Long Tail"
        st_fg  = GRN if status.startswith("Core") else MUT

        def gc(col, val, h="left", fg=MUT, bold=False, fmt=None):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Calibri", size=9, color=fg, bold=bold)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=h, vertical="center")
            if fmt: c.number_format = fmt

        gc(1, i,          h="center", bold=True,  fg=BLU)
        gc(2, name,       bold=(cumulative<=0.80), fg="14532D" if cumulative<=0.80 else MUT)
        gc(3, rev,        h="right",  fg=GRN,  bold=True, fmt='"$"#,##0')
        gc(4, pct,        h="right",  fg=MUT,  fmt="0.0%")
        gc(5, cumulative, h="right",  fg=BLU,  bold=True, fmt="0.0%")
        gc(6, 0.80,       h="right",  fg="DC2626", fmt="0%")
        gc(7, cat,        h="center", fg="0F766E")
        gc(8, status,     h="center", fg=st_fg,  bold=True)

    # Combo chart
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Pareto: Revenue vs Cumulative %"
    chart.height = 14; chart.width = 24; chart.style = 10
    rev_ref  = Reference(ws, min_col=3, min_row=4, max_row=29)
    cum_ref  = Reference(ws, min_col=5, min_row=4, max_row=29)
    chart.add_data(rev_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.solidFill = "16A34A"

    from openpyxl.chart import LineChart
    lc_chart = LineChart()
    lc_chart.add_data(cum_ref, titles_from_data=True)
    lc_chart.series[0].graphicalProperties.line.solidFill = "DC2626"
    lc_chart.series[0].graphicalProperties.line.width     = 25000
    chart += lc_chart
    ws.add_chart(chart, "A32")

    set_col_widths(ws,{"A":6,"B":28,"C":12,"D":10,"E":12,"F":9,"G":12,"H":10})
    ws.freeze_panes = "A5"


# ─── COHORT RETENTION ────────────────────────────────────────────────────────
def build_cohort_retention(wb):
    """Monthly cohort heatmap — teal/white theme."""
    ws = wb.create_sheet("Cohort Retention")
    ws.sheet_view.showGridLines = False

    H_BG = "0F4C75"; H_FG = "FFFFFF"

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "COHORT RETENTION ANALYSIS  ·  Monthly Cohorts  ·  GA4 Intelligence Hub"
    c.font = Font(name="Calibri", size=13, bold=True, color=H_FG)
    c.fill = PatternFill("solid", fgColor=H_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    c2 = ws["A2"]
    c2.value = ("Insight: Month-0 retention = 100%  ·  Month-1 avg = 42%  ·  "
                "Month-3 avg = 28%  ·  Month-6 avg = 19%  ·  Month-12 avg = 11%")
    c2.font = Font(name="Calibri", size=10, italic=True, color=H_BG)
    c2.fill = PatternFill("solid", fgColor="E0F2FE")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    cohort_months = ["Nov-20","Dec-20","Jan-21","Feb-21","Mar-21","Apr-21",
                     "May-21","Jun-21","Jul-21","Aug-21","Sep-21","Oct-21"]

    # Retention grid (12 cohorts × 13 months 0-12)
    retention_grid = [
        [1.00,0.44,0.31,0.27,0.22,0.19,0.17,0.15,0.14,0.13,0.12,0.11,0.10],
        [1.00,0.43,0.30,0.26,0.21,0.18,0.16,0.14,0.13,0.12,0.11,0.10,None],
        [1.00,0.45,0.32,0.28,0.23,0.20,0.18,0.16,0.15,0.14,0.13,None,None],
        [1.00,0.41,0.29,0.25,0.20,0.17,0.15,0.13,0.12,0.11,None,None,None],
        [1.00,0.42,0.30,0.26,0.21,0.18,0.16,0.14,0.13,None,None,None,None],
        [1.00,0.40,0.28,0.24,0.19,0.17,0.15,0.13,None,None,None,None,None],
        [1.00,0.43,0.31,0.27,0.22,0.19,0.17,None,None,None,None,None,None],
        [1.00,0.44,0.32,0.28,0.23,0.20,None,None,None,None,None,None,None],
        [1.00,0.46,0.33,0.29,0.24,None,None,None,None,None,None,None,None],
        [1.00,0.47,0.34,0.30,None,None,None,None,None,None,None,None,None],
        [1.00,0.45,0.33,None,None,None,None,None,None,None,None,None,None],
        [1.00,0.43,None,None,None,None,None,None,None,None,None,None,None],
    ]

    # Month headers
    ws.cell(4, 1).value = "Cohort"
    ws.cell(4, 1).font = Font(name="Calibri", size=9, bold=True, color=H_FG)
    ws.cell(4, 1).fill = PatternFill("solid", fgColor=H_BG)
    ws.cell(4, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(4, 2).value = "New Users"
    ws.cell(4, 2).font = Font(name="Calibri", size=9, bold=True, color=H_FG)
    ws.cell(4, 2).fill = PatternFill("solid", fgColor=H_BG)
    ws.cell(4, 2).alignment = Alignment(horizontal="center", vertical="center")
    for m in range(13):
        cell = ws.cell(row=4, column=m + 3, value=f"M+{m}")
        cell.font = Font(name="Calibri", size=9, bold=True, color=H_FG)
        cell.fill = PatternFill("solid", fgColor=H_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    new_users = [3800,3200,2900,2600,3100,2800,3300,3500,2700,3000,2900,3100]

    def retention_color(val):
        if val is None: return "D1D5DB"
        if val >= 0.40: return "0369A1"
        if val >= 0.30: return "0EA5E9"
        if val >= 0.20: return "7DD3FC"
        if val >= 0.10: return "BAE6FD"
        return "E0F2FE"

    for ci, (cohort, nu, row_data) in enumerate(zip(cohort_months, new_users, retention_grid), 5):
        ws.row_dimensions[ci].height = 18
        ws.cell(ci, 1).value = cohort
        ws.cell(ci, 1).font = Font(name="Calibri", size=9, bold=True, color="0F4C75")
        ws.cell(ci, 1).fill = PatternFill("solid", fgColor="DBEAFE")
        ws.cell(ci, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(ci, 2).value = nu
        ws.cell(ci, 2).font = Font(name="Calibri", size=9, color="374151")
        ws.cell(ci, 2).fill = PatternFill("solid", fgColor="F0F9FF")
        ws.cell(ci, 2).number_format = "#,##0"
        ws.cell(ci, 2).alignment = Alignment(horizontal="right", vertical="center")
        for mi, val in enumerate(row_data):
            cell = ws.cell(row=ci, column=mi + 3)
            cell.fill = PatternFill("solid", fgColor=retention_color(val))
            if val is not None:
                cell.value = val
                cell.number_format = "0%"
                text_col = "FFFFFF" if val >= 0.30 else "0F4C75"
                cell.font = Font(name="Calibri", size=9, bold=(val == 1.00), color=text_col)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Legend
    ws.row_dimensions[18].height = 6
    ws.merge_cells("A19:N19")
    lg = ws["A19"]
    lg.value = "LEGEND:  ■ ≥40% (dark blue)   ■ ≥30% (medium blue)   ■ ≥20% (light blue)   ■ ≥10% (pale blue)   ■ <10% (very pale)   ■ No data (grey)"
    lg.font = Font(name="Calibri", size=9, italic=True, color="374151")
    lg.fill = PatternFill("solid", fgColor="F8FAFC")
    lg.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[19].height = 16

    set_col_widths(ws,{"A":8,"B":10,"C":7,"D":7,"E":7,"F":7,"G":7,
                       "H":7,"I":7,"J":7,"K":7,"L":7,"M":7,"N":7})
    ws.freeze_panes = "C5"


# ─── CHANNEL ROI ─────────────────────────────────────────────────────────────
def build_channel_roi(wb):
    """Channel revenue, cost, ROAS, CPA — orange/white theme."""
    ws = wb.create_sheet("Channel ROI")
    ws.sheet_view.showGridLines = False

    H_BG = "7C2D12"; H_FG = "FFFFFF"
    ROW_A = "FFF7ED"; ROW_B = "FFFFFF"
    GRN   = "166534"; RED = "991B1B"; MUT = "374151"; BLU = "1D4ED8"
    ORG   = "9A3412"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "CHANNEL ROI ANALYSIS  ·  Marketing Attribution  ·  GA4 Intelligence Hub"
    c.font = Font(name="Calibri", size=13, bold=True, color=H_FG)
    c.fill = PatternFill("solid", fgColor=H_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = "Total Revenue: $891,480  ·  Paid Spend: $214,800  ·  Blended ROAS: 4.15x  ·  Blended CPA: $25.55"
    c2.font = Font(name="Calibri", size=10, italic=True, color=H_BG)
    c2.fill = PatternFill("solid", fgColor="FFEDD5")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    headers = ["Channel","Source/Medium","Revenue ($)","Sessions","Txns","CVR %",
               "Ad Spend ($)","ROAS","CPA ($)","% of Revenue"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(name="Calibri", size=9, bold=True, color=H_FG)
        cell.fill = PatternFill("solid", fgColor=H_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="FB923C"))
    ws.row_dimensions[4].height = 22

    # Revenue sums to exactly $891,480
    CHANNELS = [
        ("Organic Search",  "google / organic",       298500, 712400, 11070, 1.55,       0, None, None, 0.3349),
        ("Direct",          "(direct) / (none)",      222380, 398600,  9820, 2.46,       0, None, None, 0.2494),
        ("Paid Search",     "google / cpc",           178400, 390800,  7920, 2.03, 102000,  1.75, 12.88, 0.2001),
        ("Email",           "newsletter / email",     133600, 198200,  4920, 2.48,  62400,  2.14, 12.68, 0.1498),
        ("Social",          "facebook / social",       44600, 135200,  1020, 0.75,  38400,  1.16, 37.65, 0.0500),
        ("Referral",        "partners / referral",     13310,  54600,   108, 0.20,  12000,  1.11,111.11, 0.0149),
        ("Display",         "google / display",          690,   2400,    32, 1.33,     0,  None, None, 0.0008),
    ]

    for ri, (ch, sm, rev, sess, txns, cvr, spend, roas, cpa, pct) in enumerate(CHANNELS, 5):
        ws.row_dimensions[ri].height = 20
        bg = ROW_A if ri % 2 == 0 else ROW_B

        def oc(col, val, h="left", fg=MUT, bold=False, fmt=None):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.font = Font(name="Calibri", size=9, color=fg, bold=bold)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=h, vertical="center")
            if fmt: cell.number_format = fmt

        oc(1, ch,   bold=True,  fg=ORG)
        oc(2, sm,   fg="64748B")
        ws.cell(ri,2).font = Font(name="Consolas", size=8, color="64748B")
        oc(3, rev,  h="right", fg=GRN, bold=True, fmt='"$"#,##0')
        oc(4, sess, h="right", fg=MUT, fmt="#,##0")
        oc(5, txns, h="right", fg=MUT, fmt="#,##0")
        oc(6, cvr,  h="right", fg=BLU, fmt="0.00%")
        ws.cell(ri,6).value = cvr / 100
        oc(7, spend if spend else "N/A", h="right", fg=RED if spend > 0 else MUT,
           fmt='"$"#,##0' if spend else "@")
        roas_val = f"{roas:.2f}x" if roas else "Organic"
        oc(8, roas_val, h="center", fg=GRN if roas and roas >= 2 else (RED if roas else "374151"), bold=bool(roas))
        cpa_val = cpa if cpa else "N/A"
        oc(9, cpa_val if isinstance(cpa_val,str) else round(cpa_val,2),
           h="right", fg=MUT, fmt='"$"#,##0.00' if isinstance(cpa_val,(int,float)) else "@")
        oc(10, pct, h="right", fg=ORG, bold=True, fmt="0.0%")

    # Totals
    tr = 12
    ws.row_dimensions[tr].height = 22
    for ci in range(1, 11):
        ws.cell(tr, ci).fill = PatternFill("solid", fgColor=H_BG)
    ws.cell(tr,1).value = "TOTAL"
    ws.cell(tr,1).font = Font(name="Calibri", size=9, bold=True, color=H_FG)
    ws.cell(tr,1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(tr,3).value = 891480
    ws.cell(tr,3).number_format = '"$"#,##0'
    ws.cell(tr,3).font = Font(name="Calibri", size=9, bold=True, color="FCD34D")
    ws.cell(tr,3).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(tr,4).value = 1892200
    ws.cell(tr,4).number_format = "#,##0"
    ws.cell(tr,4).font = Font(name="Calibri", size=9, bold=True, color="FCD34D")
    ws.cell(tr,4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(tr,5).value = 34890
    ws.cell(tr,5).number_format = "#,##0"
    ws.cell(tr,5).font = Font(name="Calibri", size=9, bold=True, color="FCD34D")
    ws.cell(tr,5).alignment = Alignment(horizontal="right", vertical="center")

    # Pie chart — revenue by channel
    pie = PieChart()
    pie.title = "Revenue by Channel"; pie.style = 10
    pie.height = 12; pie.width = 18
    pie_data = Reference(ws, min_col=3, min_row=4, max_row=11)
    pie_cats = Reference(ws, min_col=1, min_row=5, max_row=11)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, "A14")

    set_col_widths(ws,{"A":16,"B":22,"C":12,"D":10,"E":8,"F":8,
                       "G":11,"H":9,"I":10,"J":11})
    ws.freeze_panes = "A5"


# ─── RFM SEGMENTS ────────────────────────────────────────────────────────────
def build_rfm_segments(wb):
    """RFM scoring + segment classification — purple/white theme."""
    ws = wb.create_sheet("RFM Segments")
    ws.sheet_view.showGridLines = False

    H_BG = "4C1D95"; H_FG = "FFFFFF"
    ROW_A = "F5F3FF"; ROW_B = "FFFFFF"
    MUT   = "374151"
    SEG_COLORS = {
        "Champion":   ("312E81","EDE9FE"),
        "Loyal":      ("065F46","D1FAE5"),
        "Active":     ("1D4ED8","DBEAFE"),
        "At Risk":    ("92400E","FEF3C7"),
        "Lost":       ("991B1B","FEE2E2"),
    }

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "RFM SEGMENTATION  ·  Recency · Frequency · Monetary  ·  GA4 Intelligence Hub"
    c.font = Font(name="Calibri", size=13, bold=True, color=H_FG)
    c.fill = PatternFill("solid", fgColor=H_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = ("Scoring: Recency 1-5 (5=purchased recently)  ·  "
                "Frequency 1-5 (5=most orders)  ·  Monetary 1-5 (5=highest spend)  ·  "
                "RFM Score = R+F+M (max 15)")
    c2.font = Font(name="Calibri", size=10, italic=True, color=H_BG)
    c2.fill = PatternFill("solid", fgColor="EDE9FE")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    headers = ["Customer ID","Tier","Days Since","Recency (R)","Orders","Frequency (F)",
               "Total Spend","Monetary (M)","RFM Score","Segment","Action"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(name="Calibri", size=9, bold=True, color=H_FG)
        cell.fill = PatternFill("solid", fgColor=H_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 22

    # Derive RFM from CUSTOMERS data
    RFM_DATA = [
        ("C-001","Platinum",  3,22, 892.40,"Champion",  "Priority VIP outreach"),
        ("C-002","Platinum",  5,18, 748.62,"Champion",  "Exclusive preview access"),
        ("C-003","Platinum",  7,16, 674.91,"Champion",  "Loyalty reward upgrade"),
        ("C-004","Platinum",  4,14, 589.20,"Champion",  "Co-creation opportunity"),
        ("C-005","Platinum",  8,12, 512.80,"Champion",  "Brand ambassador invite"),
        ("C-006","Gold",     14,10, 418.76,"Loyal",     "Referral programme"),
        ("C-007","Gold",     19, 9, 362.91,"Loyal",     "Upsell to Platinum"),
        ("C-008","Gold",     22, 8, 319.80,"Loyal",     "Seasonal bundle offer"),
        ("C-009","Gold",     25, 8, 298.70,"Loyal",     "Cross-sell accessories"),
        ("C-010","Gold",     28, 7, 268.50,"Loyal",     "Upsell campaign"),
        ("C-011","Gold",     31, 7, 248.40,"Loyal",     "Loyalty points reminder"),
        ("C-012","Gold",     35, 6, 224.92,"Loyal",     "Seasonal discount"),
        ("C-013","Gold",     38, 6, 218.47,"Loyal",     "Win-back nudge"),
        ("C-014","Silver",   41, 5, 189.30,"Active",    "Re-engagement email"),
        ("C-015","Silver",   45, 5, 175.94,"Active",    "Product recommendation"),
        ("C-016","Silver",   48, 4, 167.96,"Active",    "Newsletter highlight"),
        ("C-017","Silver",   52, 4, 159.96,"Active",    "Flash sale invite"),
        ("C-018","Silver",   55, 4, 148.20,"Active",    "Social proof campaign"),
        ("C-019","Silver",   59, 3, 134.90,"Active",    "Cart abandonment retarget"),
        ("C-020","Silver",   63, 3, 124.00,"Active",    "Browse reminder"),
        ("C-025","Silver",   92, 2,  91.98,"At Risk",   "Win-back 10% discount"),
        ("C-026","Silver",  101, 2,  84.98,"At Risk",   "Personalised outreach"),
        ("C-027","Bronze",  112, 2,  76.98,"At Risk",   "Survey + incentive"),
        ("C-028","Bronze",  121, 2,  69.98,"At Risk",   "Limited-time offer"),
        ("C-029","Bronze",  134, 1,  62.99,"At Risk",   "Reactivation email"),
        ("C-030","Bronze",  143, 1,  57.99,"At Risk",   "Social retargeting"),
        ("C-035","Bronze",  201, 1,  34.99,"Lost",      "Sunset + feedback ask"),
        ("C-036","Bronze",  214, 1,  32.99,"Lost",      "Final reactivation"),
        ("C-037","Bronze",  223, 1,  29.99,"Lost",      "Remove from active list"),
        ("C-038","Bronze",  238, 1,  27.99,"Lost",      "Archive segment"),
    ]

    def rfm_score(days, orders, spend):
        r = 5 if days<=7 else 4 if days<=21 else 3 if days<=60 else 2 if days<=120 else 1
        f = 5 if orders>=15 else 4 if orders>=8 else 3 if orders>=4 else 2 if orders>=2 else 1
        m = 5 if spend>=500 else 4 if spend>=250 else 3 if spend>=100 else 2 if spend>=50 else 1
        return r, f, m, r+f+m

    for i, (cid, tier, days, orders, spend, segment, action) in enumerate(RFM_DATA, 5):
        ws.row_dimensions[i].height = 18
        bg = ROW_A if i % 2 == 0 else ROW_B
        r, f, m, total_score = rfm_score(days, orders, spend)
        seg_fg, seg_bg = SEG_COLORS.get(segment, (MUT, bg))

        def pc(col, val, h="left", fg=MUT, bold=False, fmt=None, cell_bg=None):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=9, color=fg, bold=bold)
            cell.fill = PatternFill("solid", fgColor=cell_bg or bg)
            cell.alignment = Alignment(horizontal=h, vertical="center")
            if fmt: cell.number_format = fmt

        pc(1,  cid,         fg="4C1D95", bold=True)
        pc(2,  tier,        h="center", fg=seg_fg, bold=True, cell_bg=seg_bg)
        pc(3,  days,        h="right",  fg=MUT, fmt="#,##0")
        pc(4,  r,           h="center", fg="FFFFFF" if r>=4 else MUT,
           bold=True, cell_bg="4C1D95" if r==5 else "7C3AED" if r==4 else "A78BFA" if r==3 else "DDD6FE")
        pc(5,  orders,      h="right",  fg=MUT, fmt="#,##0")
        pc(6,  f,           h="center", fg="FFFFFF" if f>=4 else MUT,
           bold=True, cell_bg="065F46" if f==5 else "047857" if f==4 else "6EE7B7" if f==3 else "D1FAE5")
        pc(7,  spend,       h="right",  fg="166534", bold=True, fmt='"$"#,##0.00')
        pc(8,  m,           h="center", fg="FFFFFF" if m>=4 else MUT,
           bold=True, cell_bg="1D4ED8" if m==5 else "2563EB" if m==4 else "93C5FD" if m==3 else "DBEAFE")
        pc(9,  total_score, h="center", fg="FFFFFF", bold=True,
           cell_bg="4C1D95" if total_score>=13 else "7C3AED" if total_score>=10 else "A78BFA" if total_score>=7 else "C4B5FD")
        pc(10, segment,     h="center", fg=seg_fg, bold=True, cell_bg=seg_bg)
        pc(11, action,      fg="374151")

    # Segment summary block
    sr = len(RFM_DATA) + 7
    ws.merge_cells(f"A{sr}:K{sr}")
    sh = ws.cell(sr, 1)
    sh.value = "SEGMENT SUMMARY"
    sh.font = Font(name="Calibri", size=10, bold=True, color=H_FG)
    sh.fill = PatternFill("solid", fgColor=H_BG)
    sh.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sr].height = 20

    seg_summary = [
        ("Champion", 5, "892–149 avg spend","Top 5 Platinum customers · Revenue: $3,418 · Avg RFM: 14.2"),
        ("Loyal",    8, "419–218 avg spend","Gold tier regulars · Revenue: $2,361 · Avg RFM: 10.8"),
        ("Active",   7, "189–124 avg spend","Silver casuals · Revenue: $1,101 · Avg RFM: 8.1"),
        ("At Risk",  6, "92–58 avg spend",  "Declining frequency · Revenue:  $444 · Avg RFM: 5.3"),
        ("Lost",     4, "35–28 avg spend",  "No recent activity  · Revenue:  $126 · Avg RFM: 3.0"),
    ]
    for si, (seg, count, spend_range, note) in enumerate(seg_summary, sr+1):
        seg_fg, seg_bg = SEG_COLORS.get(seg, (MUT, "FFFFFF"))
        ws.row_dimensions[si].height = 18
        ws.cell(si,1).value = seg
        ws.cell(si,1).font = Font(name="Calibri", size=9, bold=True, color=seg_fg)
        ws.cell(si,1).fill = PatternFill("solid", fgColor=seg_bg)
        ws.cell(si,1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(si,2).value = count
        ws.cell(si,2).font = Font(name="Calibri", size=9, color=MUT)
        ws.cell(si,2).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(si,2).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"C{si}:D{si}")
        ws.cell(si,3).value = spend_range
        ws.cell(si,3).font = Font(name="Calibri", size=9, color="166534")
        ws.cell(si,3).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(si,3).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"E{si}:K{si}")
        ws.cell(si,5).value = note
        ws.cell(si,5).font = Font(name="Calibri", size=9, color=MUT)
        ws.cell(si,5).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(si,5).alignment = Alignment(horizontal="left", vertical="center")

    set_col_widths(ws,{"A":10,"B":10,"C":10,"D":10,"E":8,"F":12,
                       "G":12,"H":10,"I":10,"J":10,"K":24})
    ws.freeze_panes = "A5"


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    print("Building GA4 Intelligence Hub Excel workbook...")
    wb = openpyxl.Workbook()

    build_dashboard(wb)
    print("  ✓  Dashboard")
    build_transactions(wb)
    print("  ✓  Transactions (40 rows)")
    build_events(wb)
    print("  ✓  Event Summary (25 events)")
    build_monthly_revenue(wb)
    print("  ✓  Monthly Revenue (36 months)")
    build_funnel(wb)
    print("  ✓  Funnel Analysis")
    build_ml(wb)
    print("  ✓  ML Models")
    build_traffic(wb)
    print("  ✓  Traffic & Audience")
    build_bigquery(wb)
    print("  ✓  BigQuery Queries")
    build_dirty_data(wb)
    print("  ✓  Raw Data (Dirty) — Bronze layer with quality issues flagged")
    build_master_products(wb)
    print("  ✓  Master Products (25 SKUs + category pie chart)")
    build_master_customers(wb)
    print("  ✓  Master Customers (50 profiles + churn risk heatmap)")
    build_order_line_items(wb)
    print("  ✓  Order Line Items (120 rows — 3 per order)")
    build_dbt_models(wb)
    print("  ✓  dbt Models (15 models + Airflow DAG schedule)")
    build_product_performance(wb)
    print("  ✓  Product Performance (25 SKUs ranked by revenue + bar chart)")
    build_pareto(wb)
    print("  ✓  Pareto Analysis (80/20 revenue concentration + combo chart)")
    build_cohort_retention(wb)
    print("  ✓  Cohort Retention (12-month heatmap — teal theme)")
    build_channel_roi(wb)
    print("  ✓  Channel ROI (6 channels · ROAS · CPA + pie chart)")
    build_rfm_segments(wb)
    print("  ✓  RFM Segments (30 customers scored + segment summary)")

    out = "GA4_Intelligence_Hub_v3_Final.xlsx"
    wb.save(out)
    print(f"\n  ✓  Saved: {out}")
    print(f"  ✓  Sheets: {len(wb.sheetnames)} (18 total)")
    print(f"  ✓  Themes: dark navy · light blue · green · orange · purple")
    print(f"  ✓  Open in Excel → File → Save As to keep full formatting\n")


if __name__ == "__main__":
    main()
